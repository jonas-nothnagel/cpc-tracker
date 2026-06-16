"""
Parse Panama biodiversity-expenditure data into the canonical BER JSON shape.

Source: `Tablas_adicionales_desagregacion_descripciones (1).xlsx`, sheet
`Desagregación_descrita`. This is the file the Panama BIOFIN team (Magda)
prefers because it ships three augmentations the raw transactional data
lacks: a Spanish descriptive layer per expenditure line, an English
translation column, and institutional reference URLs.

The sheet carries 500 rows in a flat hierarchy (institution → programme →
subprogramme → activity), with parent rows whose money values equal the
sum of their children. There is no explicit depth column; programme rows
are identified via numeric balance — we walk the rows for each institution
block and find the rows whose values partition the institution total.

Output: `python/data/panama-ber.json`, Mongolia-shape JSON consumed
unchanged by Step 8 (`python/src/budget_align.py` + Step 8 in
`run_analysis.py:552-629`) via the existing `derive_country_file` helper.

Pipeline (each step is a pure function for testability):
    1. Load rows from the Desagregación_descrita sheet.
    2. Split rows into per-institution blocks (institution-total marks a block).
    3. Identify programme rows in each block via numeric balance.
    4. Classify programmes as substantive vs overhead (Decision 3 — name patterns).
    5. Build pseudo-targets: substantive programmes one each, overhead
       rolled up per institution.
    6. Emit panama-ber.json and a human-readable preview .md.

Decisions backing each step live in docs/panama-financial-alignment-scoping.md.
The 2026-06-17 source pivot to Tablas is recorded there too.

Usage:
    python -m src.parse_panama_ber
"""

from __future__ import annotations

import argparse
import json
import logging
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Constants and paths
# ---------------------------------------------------------------------------

REPO_ROOT = Path(__file__).resolve().parents[2]
SOURCE_XLSX = REPO_ROOT / "data" / "Panama" / "Panama Materials" / "Financial data" / "Tablas_adicionales_desagregacion_descripciones (1).xlsx"
OUTPUT_JSON = REPO_ROOT / "python" / "data" / "panama-ber.json"
PREVIEW_MD = REPO_ROOT / "python" / "output" / "panama" / "panama-ber-pseudo-targets-preview.md"

SHEET_NAME = "Desagregación_descrita"

# Period covered by the dataset. Asserted at load; emitted into the BER JSON.
PERIOD_START = 2015
PERIOD_END = 2024
ALL_YEARS = list(range(PERIOD_START, PERIOD_END + 1))

# Currency / unit emitted into the BER JSON. The sheet header reads
# "En millones de Balboas" — values are already in millions of PAB.
CURRENCY = "PAB"
UNIT = "million"

# Column indices in Desagregación_descrita (validated against header row).
COLS = {
    "institucion": 0,
    "line": 1,
    "year_start": 2,  # 2015 lives at col 2
    "year_end": 11,   # 2024 lives at col 11
    "desc_es": 12,
    "desc_en": 13,
    "fuente_url": 14,
}

# Tolerance for numeric balance when reconstructing the hierarchy.
# Tablas money values are typically 3-6 decimal places; 0.01M PAB is
# tighter than rounding error but loose enough to absorb minor drift.
BALANCE_TOLERANCE = 0.01

# Overhead detection: Tablas drops the BD sectorial codes, so we identify
# overhead programmes purely by name pattern. The seed list comes from
# observed institutional-support programme names; extend as needed.
OVERHEAD_NAME_PATTERNS = (
    "dirección y administración general",
    "dirección general",
    "dirección superior",
    "despacho superior",
    "despacho  superior",
    "servicios administrativos",
    "administración superior",
    "administraciones regionales",
    "administración general",
)


# ---------------------------------------------------------------------------
# Row representation
# ---------------------------------------------------------------------------


@dataclass
class TablasRow:
    """One line in Desagregación_descrita after header skip."""
    row_idx: int  # 1-based position in source sheet, for tracing
    institucion: str
    line: str
    values: dict[int, float]  # year → executed M PAB
    desc_es: str
    desc_en: str
    fuente_url: str

    @property
    def total_executed(self) -> float:
        return sum(self.values.values())

    @property
    def is_institution_total(self) -> bool:
        """Institution-total rows have col 0 == col 1."""
        return self.institucion == self.line


@dataclass
class ProgrammePseudo:
    """A programme-level pseudo-target (Decision 1)."""
    institution_idx: int
    programme_idx: int
    institution_name: str
    programme_name: str
    values: dict[int, float]
    desc_es: str
    desc_en: str
    fuente_url: str
    derived_from_row_count: int
    is_overhead: bool


# ---------------------------------------------------------------------------
# Step 1: Load
# ---------------------------------------------------------------------------


def validate_headers(header_row: tuple[Any, ...]) -> None:
    """Defensive check against silent upstream column re-ordering. Year
    columns accept either int (Excel-typed cell) or string ('2015')."""
    expected = {
        0: "INSTITUCIÓN",
        1: "Institución / Programa / Subprograma / Actividad",
        12: "Descripción de la línea",
        13: "Description of expenditure line",
        14: "Fuente referencial",
    }
    for idx, expected_label in expected.items():
        actual = header_row[idx]
        if actual != expected_label:
            raise ValueError(
                f"{SHEET_NAME} column {idx} expected header {expected_label!r}, "
                f"got {actual!r}. Source schema may have changed."
            )
    # Year columns 2..11 must be 2015..2024 (either int or str).
    for col_offset, year in enumerate(ALL_YEARS):
        col_idx = COLS["year_start"] + col_offset
        actual = header_row[col_idx]
        if str(actual) != str(year):
            raise ValueError(
                f"{SHEET_NAME} column {col_idx} expected year header {year}, "
                f"got {actual!r}. Source schema may have changed."
            )


def _coerce_money(v: Any) -> float:
    if v is None or v == "":
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def load_rows(xlsx_path: Path) -> list[TablasRow]:
    """Stream the Desagregación_descrita sheet and parse rows.

    Skips the title row (R0) and the header row (R1); returns data rows
    in source order. Position is preserved because hierarchy reconstruction
    relies on row ordering."""
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        ws = wb[SHEET_NAME]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    # R0 = title ("Gasto público en biodiversidad por entidad..."), R1 = headers.
    validate_headers(rows[1])

    out: list[TablasRow] = []
    for src_idx, raw in enumerate(rows[2:], start=3):  # 1-based row numbers for trace
        institucion = (raw[COLS["institucion"]] or "")
        line = (raw[COLS["line"]] or "")
        if not institucion and not line:
            continue
        values = {
            year: _coerce_money(raw[COLS["year_start"] + (year - PERIOD_START)])
            for year in ALL_YEARS
        }
        out.append(TablasRow(
            row_idx=src_idx,
            institucion=str(institucion).strip(),
            line=str(line).strip(),
            values=values,
            desc_es=str(raw[COLS["desc_es"]] or "").strip(),
            desc_en=str(raw[COLS["desc_en"]] or "").strip(),
            fuente_url=str(raw[COLS["fuente_url"]] or "").strip(),
        ))
    return out


# ---------------------------------------------------------------------------
# Step 2: Split into institution blocks
# ---------------------------------------------------------------------------


def split_into_institution_blocks(rows: list[TablasRow]) -> list[tuple[TablasRow, list[TablasRow]]]:
    """Group rows by institution. Returns a list of (institution_total_row, child_rows)
    tuples in source order. An institution block starts at an institution-total row
    and runs until the next institution-total row or end-of-list."""
    blocks: list[tuple[TablasRow, list[TablasRow]]] = []
    current_total: TablasRow | None = None
    current_children: list[TablasRow] = []
    for row in rows:
        if row.is_institution_total:
            if current_total is not None:
                blocks.append((current_total, current_children))
            current_total = row
            current_children = []
        else:
            if current_total is None:
                # Orphan rows before first institution-total — skip with a warning
                logger.warning(f"Orphan row {row.row_idx}: {row.line[:40]!r} appears before any institution-total")
                continue
            current_children.append(row)
    if current_total is not None:
        blocks.append((current_total, current_children))
    return blocks


# ---------------------------------------------------------------------------
# Step 3: Identify programme rows via numeric balance
# ---------------------------------------------------------------------------


def _absorb_subtree(
    children: list[TablasRow],
    start_idx: int,
    parent_value: float,
) -> tuple[list[TablasRow], int]:
    """Absorb all descendants of a parent row whose value is `parent_value`.

    Walks `children` from `start_idx`, greedily including rows as direct
    children of the parent until their cumulative value reaches `parent_value`
    (within tolerance). Each direct child recursively absorbs its own
    descendants. Stops if the next row's value would overshoot — that signals
    the next row belongs at a higher level of the hierarchy.

    Returns (absorbed_rows, next_idx). The absorbed list flattens the whole
    subtree so callers don't have to recurse to count rows."""
    absorbed: list[TablasRow] = []
    accumulated = 0.0
    j = start_idx
    while j < len(children) and accumulated + BALANCE_TOLERANCE < parent_value:
        child = children[j]
        if accumulated + child.total_executed > parent_value + BALANCE_TOLERANCE:
            # Adding this child would overshoot — it belongs to a higher level.
            break
        absorbed.append(child)
        accumulated += child.total_executed
        j += 1
        # Recursively absorb this child's own descendants.
        descendant_rows, j = _absorb_subtree(children, j, child.total_executed)
        absorbed.extend(descendant_rows)
    return absorbed, j


def identify_programme_rows(
    institution_total: TablasRow,
    children: list[TablasRow],
) -> list[tuple[TablasRow, list[TablasRow]]]:
    """Walk a single institution block and partition its children into
    (programme_row, drill_down_rows) tuples.

    Tablas encodes the hierarchy via numeric balance: parent.value equals
    the sum of its direct children's values. We exploit that to identify
    programme-level rows: at each step, the current row is taken as a
    programme, then `_absorb_subtree` consumes all its descendants. Programme
    rows are accepted as long as the running programme tally hasn't yet
    reached the institution total.
    """
    target_total = institution_total.total_executed
    if not children:
        return []

    programmes: list[tuple[TablasRow, list[TablasRow]]] = []
    i = 0
    institution_running = 0.0

    while i < len(children) and institution_running + BALANCE_TOLERANCE < target_total:
        programme = children[i]
        if institution_running + programme.total_executed > target_total + BALANCE_TOLERANCE:
            # Treating this row as a programme would overshoot — must be a
            # deeper-level orphan that we should silently absorb under the
            # previous programme. Append to previous drill_down if any.
            if programmes:
                programmes[-1][1].append(programme)
                i += 1
                continue
            else:
                logger.warning(
                    f"Institution {institution_total.line!r}: first row "
                    f"{programme.line!r} (value {programme.total_executed:.4f}M) "
                    f"exceeds institution total {target_total:.4f}M. Skipping."
                )
                break
        i += 1
        drill_down, i = _absorb_subtree(children, i, programme.total_executed)
        programmes.append((programme, drill_down))
        institution_running += programme.total_executed

    if abs(institution_running - target_total) > BALANCE_TOLERANCE:
        logger.warning(
            f"Institution {institution_total.line!r}: identified programmes sum to "
            f"{institution_running:.4f}M but institution total is {target_total:.4f}M "
            f"(delta {institution_running - target_total:+.4f}M). Likely orphan rows; "
            f"{len(children) - i} rows left unattributed."
        )
    return programmes


# ---------------------------------------------------------------------------
# Step 4: Overhead classification (Decision 3, name-pattern based)
# ---------------------------------------------------------------------------


def is_overhead_name(programme_name: str) -> bool:
    """Decide whether a programme name reads as institutional overhead.
    Tablas drops the BD sectorial codes, so this is name-pattern only —
    less precise than the BD version but adequate for the recurring
    overhead-programme names."""
    lower = programme_name.lower().strip()
    return any(pattern in lower for pattern in OVERHEAD_NAME_PATTERNS)


# ---------------------------------------------------------------------------
# Step 5: Build pseudo-targets
# ---------------------------------------------------------------------------


def _round_money(v: float) -> float:
    return round(v, 4)


def _to_year_dict(values: dict[int, float]) -> dict[str, float | None]:
    """Convert {year: M PAB} to {year_str: M PAB or null}. Zero years emit 0.0;
    we never see missing years in Tablas (all year columns are populated)."""
    return {str(y): _round_money(values.get(y, 0.0)) for y in ALL_YEARS}


def _institution_name_short(name: str) -> str:
    """Title-case a SHOUTY institution name for use in description text."""
    return name.title()


def build_substantive_pseudo(
    institution_idx: int,
    programme_idx: int,
    institution_name: str,
    programme: "TablasRow",
    drill_down: list["TablasRow"],
) -> ProgrammePseudo:
    """Build a substantive (non-overhead) programme pseudo-target. Description
    text is rendered at emit time via _render_description so the same logic
    serves both panama-ber.json and the preview .md."""
    return ProgrammePseudo(
        institution_idx=institution_idx,
        programme_idx=programme_idx,
        institution_name=institution_name,
        programme_name=programme.line,
        values=dict(programme.values),
        desc_es=programme.desc_es,
        desc_en=programme.desc_en,
        fuente_url=programme.fuente_url,
        derived_from_row_count=1 + len(drill_down),
        is_overhead=False,
    )


def build_overhead_rollup_pseudo(
    institution_idx: int,
    institution_name: str,
    overhead_programmes: list[tuple["TablasRow", list["TablasRow"]]],
) -> ProgrammePseudo:
    """Roll up all overhead programmes for an institution into one pseudo-target."""
    combined_values: dict[int, float] = {y: 0.0 for y in ALL_YEARS}
    row_count = 0
    programme_names: list[str] = []
    descs_es: list[str] = []
    descs_en: list[str] = []
    urls: set[str] = set()
    for programme, drill_down in overhead_programmes:
        for y, v in programme.values.items():
            combined_values[y] += v
        row_count += 1 + len(drill_down)
        programme_names.append(programme.line)
        if programme.desc_es:
            descs_es.append(programme.desc_es)
        if programme.desc_en:
            descs_en.append(programme.desc_en)
        if programme.fuente_url:
            urls.add(programme.fuente_url)

    # Pick the most informative EN description (longest), fall back to ES.
    # Stored on the pseudo so _render_description can use it at emit time.
    desc_en_pick = max(descs_en, key=len) if descs_en else ""
    desc_es_pick = max(descs_es, key=len) if descs_es else ""

    return ProgrammePseudo(
        institution_idx=institution_idx,
        programme_idx=0,  # 0 marks overhead rollup
        institution_name=institution_name,
        programme_name=f"Apoyo institucional — {_institution_name_short(institution_name)}",
        values=combined_values,
        desc_es=desc_es_pick,
        desc_en=desc_en_pick,
        fuente_url=" ; ".join(sorted(urls)),
        derived_from_row_count=row_count,
        is_overhead=True,
    )


# ---------------------------------------------------------------------------
# Step 6: Emit
# ---------------------------------------------------------------------------


def _render_description(p: ProgrammePseudo) -> str:
    """Render the pseudo-target's description text. Substantive vs overhead
    differ in framing; both quote the Tablas LLM description verbatim and
    state Tablas as the descriptive source."""
    total = sum(p.values.values())
    avg = total / len(ALL_YEARS) if ALL_YEARS else 0.0
    desc_text = p.desc_en or p.desc_es or "(no description available)"
    if p.is_overhead:
        return (
            f"Cross-cutting institutional support and general administration at the "
            f"{p.institution_name}. {desc_text} "
            f"Executed expenditure {PERIOD_START}–{PERIOD_END}: "
            f"{total:.1f}M PAB total (avg {avg:.1f}M/year). "
            f"Descriptions sourced from Tablas_adicionales descriptive layer (LLM-generated, Panama BIOFIN team preferred)."
        )
    return (
        f'Programme "{p.programme_name}" under the {p.institution_name}. '
        f"{desc_text} "
        f"Executed expenditure {PERIOD_START}–{PERIOD_END}: "
        f"{total:.1f}M PAB total (avg {avg:.1f}M/year). "
        f"Description sourced from Tablas_adicionales descriptive layer (LLM-generated, Panama BIOFIN team preferred)."
    )


def assemble_ber_payload(pseudos: list[ProgrammePseudo]) -> dict[str, Any]:
    """Assemble final BER JSON payload (Mongolia-shape)."""
    programs: list[dict[str, Any]] = []
    expenditure: list[dict[str, Any]] = []
    for p in pseudos:
        if p.is_overhead:
            code = f"BER_PA_{p.institution_idx:02d}_OVERHEAD"
        else:
            code = f"BER_PA_{p.institution_idx:02d}_{p.programme_idx:02d}"
        programs.append({
            "code": code,
            "name": p.programme_name,
            "description": _render_description(p),
            "type": "environmental",
        })
        expenditure.append({
            "code": code,
            "name": p.programme_name,
            "values": _to_year_dict(p.values),
        })
    return {
        "programs": programs,
        "expenditure": expenditure,
        "currency": CURRENCY,
        "unit": UNIT,
        "period": {"start": PERIOD_START, "end": PERIOD_END},
    }


def render_preview_md(pseudos: list[ProgrammePseudo]) -> str:
    """HITL preview .md (Decision 6) — every pseudo-target with its
    LLM-input description, source URL, and row count."""
    lines: list[str] = []
    lines.append("# Panama BER pseudo-targets — preview")
    lines.append("")
    lines.append(
        f"Generated by `python -m src.parse_panama_ber`. "
        f"Source: `data/Panama/Panama Materials/Financial data/Tablas_adicionales_desagregacion_descripciones (1).xlsx`, "
        f"sheet `{SHEET_NAME}`. Period {PERIOD_START}-{PERIOD_END}. Money in millions of {CURRENCY}. "
        f"Descriptions are from Tablas' LLM-generated descriptive layer — preferred by the Panama BIOFIN team for interpretability."
    )
    lines.append("")
    lines.append(
        "Review the descriptions below before triggering Step 8 of `run_analysis.py`. "
        "Each description is what the alignment LLM sees per programme. "
        "Look for: opaque programme names, missing English translations, and overhead rollups that should/shouldn't be rolled up."
    )
    lines.append("")
    substantive = [p for p in pseudos if not p.is_overhead]
    overheads = [p for p in pseudos if p.is_overhead]
    lines.append(
        f"## {len(pseudos)} pseudo-targets — {len(substantive)} substantive, {len(overheads)} institutional overhead rollups"
    )
    lines.append("")
    for p in pseudos:
        code = (
            f"BER_PA_{p.institution_idx:02d}_OVERHEAD"
            if p.is_overhead else
            f"BER_PA_{p.institution_idx:02d}_{p.programme_idx:02d}"
        )
        lines.append(f"### {code} — {p.programme_name}")
        lines.append("")
        lines.append("**Description (LLM input)**:")
        lines.append("")
        lines.append(f"> {_render_description(p)}")
        lines.append("")
        # Per-year executed values.
        yearline = " · ".join(
            f"{y}: {_round_money(p.values.get(y, 0.0)):.2f}M" for y in ALL_YEARS
        )
        lines.append(f"**Per year (M PAB)**: {yearline}")
        lines.append("")
        lines.append(f"**Derived from**: {p.derived_from_row_count} source rows in `{SHEET_NAME}`")
        if p.fuente_url:
            lines.append(f"**Reference**: {p.fuente_url}")
        lines.append("")
        lines.append("---")
        lines.append("")
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Orchestrator
# ---------------------------------------------------------------------------


def run_ingest(source_xlsx: Path) -> tuple[dict[str, Any], str]:
    """End-to-end ingest. Returns (ber_payload, preview_md)."""
    logger.info(f"Reading rows from {source_xlsx}")
    rows = load_rows(source_xlsx)
    logger.info(f"Loaded {len(rows)} data rows from {SHEET_NAME}")

    blocks = split_into_institution_blocks(rows)
    logger.info(f"Split into {len(blocks)} institution blocks")

    pseudos: list[ProgrammePseudo] = []
    for inst_idx, (inst_total, children) in enumerate(blocks, start=1):
        programmes = identify_programme_rows(inst_total, children)
        substantive: list[tuple[TablasRow, list[TablasRow]]] = []
        overheads: list[tuple[TablasRow, list[TablasRow]]] = []
        for programme, drill_down in programmes:
            if is_overhead_name(programme.line):
                overheads.append((programme, drill_down))
            else:
                substantive.append((programme, drill_down))

        for prog_idx, (programme, drill_down) in enumerate(substantive, start=1):
            pseudos.append(build_substantive_pseudo(
                inst_idx, prog_idx, inst_total.line, programme, drill_down,
            ))
        if overheads:
            pseudos.append(build_overhead_rollup_pseudo(
                inst_idx, inst_total.line, overheads,
            ))

    payload = assemble_ber_payload(pseudos)
    preview = render_preview_md(pseudos)
    logger.info(
        f"Emitted {len(payload['programs'])} pseudo-targets "
        f"({sum(1 for p in pseudos if not p.is_overhead)} substantive + "
        f"{sum(1 for p in pseudos if p.is_overhead)} overhead rollups)"
    )
    return payload, preview


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--source",
        default=str(SOURCE_XLSX),
        help=f"Path to Tablas_adicionales_desagregacion_descripciones (1).xlsx (default: {SOURCE_XLSX})",
    )
    parser.add_argument(
        "--output",
        default=str(OUTPUT_JSON),
        help=f"Output JSON path (default: {OUTPUT_JSON})",
    )
    parser.add_argument(
        "--preview-output",
        default=str(PREVIEW_MD),
        help=f"Preview markdown path (default: {PREVIEW_MD})",
    )
    parser.add_argument("-v", "--verbose", action="store_true")
    args = parser.parse_args(argv)

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(levelname)s %(message)s",
    )

    source = Path(args.source)
    if not source.exists():
        logger.error(f"Source file not found: {source}")
        return 2

    payload, preview = run_ingest(source)

    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(payload, indent=2, ensure_ascii=False))
    logger.info(f"Wrote {output_path} ({len(payload['programs'])} programmes)")

    preview_path = Path(args.preview_output)
    preview_path.parent.mkdir(parents=True, exist_ok=True)
    preview_path.write_text(preview)
    logger.info(f"Wrote preview {preview_path}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
