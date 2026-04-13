"""
Generic CTF (Common Tabular Format) parser for BTR Excel files.

Supports:
  - CTF-NDC: Tables 4 (progress indicators), 5 (mitigation measures),
             6 (sector emissions), 7/9 (projections)
  - CTF-FTC-Support: Tables 9 (technology support), 11 (capacity building)

Validated against Mongolia, Lebanon, and Panama CTF files.
"""

from __future__ import annotations

import json
import logging
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)

NOTATION_KEYS = {"NA", "NE", "NO", "IE", "FX", "UA", "C"}


def _is_notation(val: Any) -> bool:
    if val is None:
        return True
    s = str(val).strip()
    if not s:
        return True
    parts = {p.strip() for p in s.replace(",", " ").split()}
    return bool(parts) and parts.issubset(NOTATION_KEYS)


def _num(val: Any) -> float | None:
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(",", "")
    if _is_notation(s):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _str(val: Any) -> str:
    if val is None:
        return ""
    s = str(val).strip()
    s = s.replace("_x000D_", " ").replace("\r\n", " ").replace("\r", " ")
    s = re.sub(r"\s+", " ", s)
    return s


_FOOTNOTE_REF_RE = re.compile(r"\(\d+\)")


def _strip_footnote_refs(s: str) -> str:
    """Strip inline CTF footnote references like '(3)' or '(54)' from cell values."""
    return _FOOTNOTE_REF_RE.sub("", s).strip()


def _find_sheet(wb: Any, patterns: list[str]) -> Worksheet | None:
    for name in wb.sheetnames:
        lower = name.lower().replace(" ", "").replace("_", "")
        for pat in patterns:
            if pat in lower:
                return wb[name]
    return None


def _find_header_row(ws: Worksheet, max_row: int = 15) -> int:
    """Find the row index where column B has a header-like value."""
    for row_idx in range(1, max_row + 1):
        val = _str(ws.cell(row=row_idx, column=2).value).lower()
        if any(kw in val for kw in ["indicator", "name", "category", "title"]):
            return row_idx
    return 7


# ---------------------------------------------------------------------------
# Table 4: Progress Indicators
# ---------------------------------------------------------------------------

def _parse_table4(ws: Worksheet) -> list[dict[str, Any]]:
    """Parse Table 4.x progress indicators."""
    indicators: list[dict[str, Any]] = []

    header_row = _find_header_row(ws)

    year_cols: dict[str, int] = {}
    for col_idx in range(2, ws.max_column + 1):
        h1 = _str(ws.cell(row=header_row - 1, column=col_idx).value)
        h2 = _str(ws.cell(row=header_row, column=col_idx).value)
        header = f"{h1} {h2}".strip().lower()

        year_match = re.search(r"\b(19|20)\d{2}\b", header)
        if year_match:
            year = year_match.group()
            if year not in year_cols:
                year_cols[year] = col_idx

    for row_idx in range(header_row + 1, min(header_row + 30, ws.max_row + 1)):
        name = _str(ws.cell(row=row_idx, column=2).value)
        if not name or name.lower().startswith("footnote"):
            continue

        unit = _str(ws.cell(row=row_idx, column=3).value)

        yearly: dict[str, float | None] = {}
        for year_str, col in year_cols.items():
            yearly[year_str] = _num(ws.cell(row=row_idx, column=col).value)

        target_level = None
        target_year = None
        progress_text = ""
        for col_idx in range(2, ws.max_column + 1):
            h = _str(ws.cell(row=header_row, column=col_idx).value).lower()
            if "target level" in h or "target" in h and "year" not in h:
                target_level = _num(ws.cell(row=row_idx, column=col_idx).value)
            elif "target year" in h or "target" in h and "period" in h:
                v = _str(ws.cell(row=row_idx, column=col_idx).value)
                target_year = v if v else None
            elif "progress" in h or "assessment" in h:
                progress_text = _str(ws.cell(row=row_idx, column=col_idx).value)

        if any(v is not None for v in yearly.values()) or target_level is not None:
            indicators.append({
                "name": name,
                "unit": unit,
                "yearlyValues": {k: v for k, v in yearly.items() if v is not None},
                "targetLevel": target_level,
                "targetYear": target_year,
                "progressText": progress_text,
            })

    return indicators


# ---------------------------------------------------------------------------
# Table 5: Mitigation Policies and Measures
# ---------------------------------------------------------------------------

def _parse_table5(ws: Worksheet) -> list[dict[str, Any]]:
    """Parse Table 5 mitigation policies/measures."""
    measures: list[dict[str, Any]] = []

    header_row = _find_header_row(ws)

    col_map: dict[str, int] = {}
    for col_idx in range(2, ws.max_column + 1):
        h = _str(ws.cell(row=header_row, column=col_idx).value).lower()
        h = re.sub(r"[a-z],?\s*$", "", h).strip()
        if "name" in h and "col" not in h:
            col_map["name"] = col_idx
        elif "description" in h:
            col_map["description"] = col_idx
        elif "objective" in h:
            col_map["objectives"] = col_idx
        elif "type of instrument" in h or "instrument" in h:
            col_map["instrumentType"] = col_idx
        elif "status" in h:
            col_map["status"] = col_idx
        elif "sector" in h:
            col_map["sector"] = col_idx
        elif "gas" in h and "affect" in h:
            col_map["gases"] = col_idx
        elif "start year" in h:
            col_map["startYear"] = col_idx
        elif "implementing" in h or "entity" in h:
            col_map["implementingEntity"] = col_idx

    estimate_cols: dict[str, int] = {}
    for col_idx in range(2, ws.max_column + 1):
        h1 = _str(ws.cell(row=header_row - 1, column=col_idx).value)
        h2 = _str(ws.cell(row=header_row, column=col_idx).value)
        combined = f"{h1} {h2}".strip().lower()
        if "estimate" in combined or "reduction" in combined or "achieved" in combined or "expected" in combined:
            year_match = re.search(r"\b(20\d{2})\b", combined)
            if year_match:
                estimate_cols[year_match.group()] = col_idx

    _footnote_re_t5 = re.compile(r"^[a-z]\s+[A-Z]")

    for row_idx in range(header_row + 1, ws.max_row + 1):
        name = _str(ws.cell(row=row_idx, column=col_map.get("name", 2)).value)
        if not name or name.lower().startswith("footnote") or len(name) < 3:
            continue
        if _footnote_re_t5.match(name):
            continue
        status = _strip_footnote_refs(_str(ws.cell(row=row_idx, column=col_map.get("status", 6)).value))
        if not status:
            continue

        estimates: dict[str, float | None] = {}
        for year_str, col in estimate_cols.items():
            estimates[year_str] = _num(ws.cell(row=row_idx, column=col).value)

        raw_sector = _strip_footnote_refs(_str(ws.cell(row=row_idx, column=col_map.get("sector", 7)).value))
        # Normalize known typo: "Sistema Marino-Costeros" → "Sistemas Marino-Costeros"
        if raw_sector == "Sistema Marino-Costeros":
            raw_sector = "Sistemas Marino-Costeros"
        measure = {
            "name": name,
            "description": _str(ws.cell(row=row_idx, column=col_map.get("description", 3)).value),
            "objectives": _str(ws.cell(row=row_idx, column=col_map.get("objectives", 4)).value),
            "instrumentType": _strip_footnote_refs(_str(ws.cell(row=row_idx, column=col_map.get("instrumentType", 5)).value)),
            "status": status,
            "sector": "",
            "sectorRaw": raw_sector,
            "gasesAffected": _str(ws.cell(row=row_idx, column=col_map.get("gases", 8)).value),
            "startYear": _strip_footnote_refs(_str(ws.cell(row=row_idx, column=col_map.get("startYear", 9)).value)),
            "implementingEntity": _str(ws.cell(row=row_idx, column=col_map.get("implementingEntity", 10)).value),
            "reductionEstimates": {k: v for k, v in estimates.items() if v is not None},
        }
        measures.append(measure)

    return measures


# ---------------------------------------------------------------------------
# Table 6: GHG Emissions by Sector
# ---------------------------------------------------------------------------

def _parse_table6(ws: Worksheet) -> dict[str, Any]:
    """Parse Table 6 sector emissions time series."""
    sector_rows: list[dict[str, Any]] = []

    year_cols: dict[str, int] = {}
    header_row = 8
    for row_idx in range(5, 12):
        val = _str(ws.cell(row=row_idx, column=2).value).lower()
        if "category" in val or "gas" in val:
            header_row = row_idx
            break

    for col_idx in range(3, ws.max_column + 1):
        for check_row in [header_row - 1, header_row]:
            h = _str(ws.cell(row=check_row, column=col_idx).value)
            year_match = re.match(r"^(19|20)\d{2}$", h.strip())
            if year_match:
                year_cols[h.strip()] = col_idx

    sector_pattern = re.compile(r"^\d+\.\s*(.+)$")
    total_pattern = re.compile(r"^total\s*\(", re.IGNORECASE)

    for row_idx in range(header_row + 1, ws.max_row + 1):
        label = _str(ws.cell(row=row_idx, column=2).value)
        if not label:
            continue

        m = sector_pattern.match(label)
        is_total = bool(total_pattern.match(label))

        if not m and not is_total:
            continue

        sector_name = m.group(1).strip() if m else label.strip()

        yearly: dict[str, float | None] = {}
        for year_str, col in year_cols.items():
            yearly[year_str] = _num(ws.cell(row=row_idx, column=col).value)

        if any(v is not None for v in yearly.values()):
            sector_rows.append({
                "category": sector_name,
                "normalizedSector": "" if is_total else _normalize_sector(sector_name),
                "isTotal": is_total,
                "yearlyEmissions": {k: v for k, v in yearly.items() if v is not None},
                "unit": "kt CO2 eq",
            })

    return {"bySector": sector_rows}


# ---------------------------------------------------------------------------
# Table 7/9: Projections
# ---------------------------------------------------------------------------

def _parse_projections(ws: Worksheet, scenario: str) -> list[dict[str, Any]]:
    """Parse projection tables (7=WEM, 8=WOM, 9=WAM)."""
    projections: list[dict[str, Any]] = []

    header_row = 8
    for row_idx in range(5, 12):
        val = _str(ws.cell(row=row_idx, column=2).value).lower()
        if "category" in val or "gas" in val:
            header_row = row_idx
            break

    year_cols: dict[str, int] = {}
    for col_idx in range(3, ws.max_column + 1):
        for check_row in [header_row - 1, header_row]:
            h = _str(ws.cell(row=check_row, column=col_idx).value)
            year_match = re.match(r"^(19|20)\d{2}$", h.strip())
            if year_match:
                year_cols[h.strip()] = col_idx

    sector_pattern = re.compile(r"^\d+\.\s*(.+)$")
    total_pattern = re.compile(r"^total\s*\(", re.IGNORECASE)
    named_sectors = {"energy", "transport", "industrial", "agriculture", "forestry",
                     "lulucf", "waste", "land use"}

    for row_idx in range(header_row + 1, ws.max_row + 1):
        label = _str(ws.cell(row=row_idx, column=2).value)
        if not label:
            continue

        m = sector_pattern.match(label)
        is_total = bool(total_pattern.match(label))
        is_named = any(kw in label.lower() for kw in named_sectors)

        if not m and not is_total and not is_named:
            continue

        sector_name = m.group(1).strip() if m else label.strip()

        yearly: dict[str, float | None] = {}
        for year_str, col in year_cols.items():
            yearly[year_str] = _num(ws.cell(row=row_idx, column=col).value)

        if any(v is not None for v in yearly.values()):
            projections.append({
                "scenario": scenario,
                "category": sector_name,
                "normalizedSector": "" if is_total else _normalize_sector(sector_name),
                "isTotal": is_total,
                "yearlyValues": {k: v for k, v in yearly.items() if v is not None},
                "unit": "kt CO2 eq",
            })

    return projections


# ---------------------------------------------------------------------------
# CTF-FTC-Support: Tables 9 / 11
# ---------------------------------------------------------------------------

# Metadata rows that appear at the bottom of CTF support tables.
_METADATA_PATTERNS = [
    re.compile(r"^[a-z]\s+[A-Z]"),               # letter-labeled footnotes: "a  Explanation..."
    re.compile(r"^\(\d+\)"),                       # numbered footnotes: "(1) The underlying..."
    re.compile(r"^Notation key", re.IGNORECASE),   # notation key legend
    re.compile(r"^Custom footnote", re.IGNORECASE),  # section header for footnotes
]

# CTF internal table references appended to titles (e.g. "- TableIII.7")
_TABLE_REF_RE = re.compile(r"\s*-\s*Table[A-Z]+\.?\d*\s*$")


def _parse_support_table(ws: Worksheet) -> list[dict[str, Any]]:
    """Parse a support-received table (technology or capacity-building)."""
    projects: list[dict[str, Any]] = []

    header_row = 6
    for row_idx in range(4, 10):
        val = _str(ws.cell(row=row_idx, column=2).value).lower()
        if "title" in val or "activity" in val:
            header_row = row_idx
            break

    col_map: dict[str, int] = {}
    for col_idx in range(2, ws.max_column + 1):
        h = _str(ws.cell(row=header_row, column=col_idx).value).lower()
        h = re.sub(r"[a-g],?\s*$", "", h).strip()
        if "title" in h:
            col_map["title"] = col_idx
        elif "description" in h and "programme" in h:
            col_map["description"] = col_idx
        elif "type of technology" in h:
            col_map["technologyType"] = col_idx
        elif "time frame" in h or "timeframe" in h:
            col_map["timeFrame"] = col_idx
        elif "recipient" in h:
            col_map["recipientEntity"] = col_idx
        elif "implementing" in h:
            col_map["implementingEntity"] = col_idx
        elif "type of support" in h:
            col_map["supportType"] = col_idx
        elif h.startswith("sector") and "sub" not in h:
            col_map["sector"] = col_idx
        elif "subsector" in h or "sub-sector" in h:
            col_map["subsector"] = col_idx
        elif "status" in h:
            col_map["status"] = col_idx
        elif "impact" in h or "result" in h:
            col_map["impact"] = col_idx
        elif "additional" in h:
            col_map["additionalInfo"] = col_idx

    title_col = col_map.get("title", 2)

    for row_idx in range(header_row + 1, ws.max_row + 1):
        title = _str(ws.cell(row=row_idx, column=title_col).value)
        if not title or len(title) < 3:
            continue
        # Strip CTF table references (e.g. "- TableIII.7") from titles
        title = _TABLE_REF_RE.sub("", title).strip()
        if any(pat.match(title) for pat in _METADATA_PATTERNS):
            continue

        project: dict[str, Any] = {"title": title}
        for field, col in col_map.items():
            if field == "title":
                continue
            val = _str(ws.cell(row=row_idx, column=col).value)
            if field == "sector":
                project[field] = _normalize_sector(val)
                project["sectorRaw"] = val
            else:
                project[field] = val

        # Structural guard: reject rows where all substantive fields are empty
        if all(not project.get(f) for f in ("supportType", "description", "timeFrame", "implementingEntity")):
            continue

        projects.append(project)

    return projects


# ---------------------------------------------------------------------------
# Sector normalization
# ---------------------------------------------------------------------------

SECTOR_MAP: dict[str, str] = {
    # English (IPCC / UNFCCC standard)
    "energy": "sector_energy",
    "transport": "sector_transport",
    "industrial processes": "sector_ippu",
    "industrial processes and product use": "sector_ippu",
    "ippu": "sector_ippu",
    "industry": "sector_ippu",
    "manufacture": "sector_ippu",
    "agriculture": "sector_agriculture",
    "land use": "sector_lulucf",
    "land use, land-use change and forestry": "sector_lulucf",
    "lulucf": "sector_lulucf",
    "forestry": "sector_lulucf",
    "forestry/lulucf": "sector_lulucf",
    "waste": "sector_waste",
    "waste management": "sector_waste",
    "waste management/waste": "sector_waste",
    "water and sanitation": "sector_other",
    "water": "sector_other",
    "infrastructure": "sector_other",
    "cross-cutting": "sector_other",
    "other": "sector_other",
    "other (specify)": "sector_other",
    # Spanish — Panama NDC sectors (lossy IPCC mapping; sectorRaw preserves original)
    "bosques": "sector_lulucf",
    "agricultura ganaderia y acuicultura sostenible": "sector_agriculture",
    "economia circular": "sector_waste",
    "gestion integrada de cuencas hidrograficas": "sector_other",
    "sistemas marino-costeros": "sector_other",
    "sistema marino-costeros": "sector_other",
    "biodiversidad": "sector_other",
    "asentamientos humanos resilientes": "sector_other",
    "salud publica": "sector_other",
    "infraestructura sostenible": "sector_other",
    "fortalecimiento de capacidades para la accion y transparencia climatica": "sector_other",
}


def _strip_accents(s: str) -> str:
    """NFD-decompose and strip combining marks so accented text matches ASCII keys."""
    import unicodedata
    return "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")


def _normalize_sector(raw: str) -> str:
    lower = _strip_accents(raw).lower().strip()
    lower = re.sub(r"^\d+\.\s*", "", lower)
    if not lower:
        return "sector_other"

    if lower in SECTOR_MAP:
        return SECTOR_MAP[lower]

    for key, sector_id in SECTOR_MAP.items():
        if key in lower or lower in key:
            return sector_id

    return "sector_other"


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def parse_ctf_ndc(filepath: str | Path) -> dict[str, Any]:
    """
    Parse a CTF-NDC Excel file.

    Returns dict with keys:
      progressIndicators, mitigationMeasures, sectorEmissions, projections
    """
    filepath = Path(filepath)
    logger.info(f"Parsing CTF-NDC: {filepath.name}")
    wb = load_workbook(filepath, data_only=True, read_only=True)

    result: dict[str, Any] = {
        "sourceFile": filepath.name,
        "progressIndicators": [],
        "mitigationMeasures": [],
        "sectorEmissions": {"bySector": []},
        "projections": [],
    }

    # Table 4.x (may have 4.1, 4.2, 4.3 sub-sheets)
    for name in wb.sheetnames:
        if "table4" in name.lower().replace(" ", "").replace("_", ""):
            ws = wb[name]
            indicators = _parse_table4(ws)
            if indicators:
                for ind in indicators:
                    ind["sourceSheet"] = name
                result["progressIndicators"].extend(indicators)
                logger.info(f"  {name}: {len(indicators)} progress indicators")

    # Table 5
    ws5 = _find_sheet(wb, ["table5"])
    if ws5:
        measures = _parse_table5(ws5)
        result["mitigationMeasures"] = measures
        logger.info(f"  Table5: {len(measures)} mitigation measures")

    # Table 6
    ws6 = _find_sheet(wb, ["table6"])
    if ws6:
        emissions = _parse_table6(ws6)
        result["sectorEmissions"] = emissions
        sector_count = len([s for s in emissions["bySector"] if not s.get("isTotal")])
        logger.info(f"  Table6: {sector_count} sector emission series")

    # Table 7 (WEM), Table 8 (WOM), Table 9 (WAM)
    for table_name, scenario in [("table7", "wem"), ("table8", "wom"), ("table9", "wam")]:
        ws = _find_sheet(wb, [table_name])
        if ws:
            projs = _parse_projections(ws, scenario)
            if projs:
                result["projections"].extend(projs)
                logger.info(f"  {table_name}: {len(projs)} projection series ({scenario})")

    wb.close()
    return result


def _merge_support_projects(
    tech: list[dict[str, Any]], cb: list[dict[str, Any]]
) -> list[dict[str, Any]]:
    """Deduplicate projects across technology support and capacity building.

    Projects appearing in both tables are merged into a single entry with
    supportSource="both". Fields from the technology entry take precedence,
    with capacity-building fields filling in blanks.
    """
    by_key: dict[str, dict[str, Any]] = {}
    for project in tech:
        key = project["title"].strip().lower()
        by_key[key] = {**project, "supportSource": "technology"}
    for project in cb:
        key = project["title"].strip().lower()
        if key in by_key:
            # Merge: fill empty fields from CB entry
            existing = by_key[key]
            existing["supportSource"] = "both"
            for field, val in project.items():
                if field == "title":
                    continue
                if val and not existing.get(field):
                    existing[field] = val
        else:
            by_key[key] = {**project, "supportSource": "capacity_building"}
    return list(by_key.values())


def parse_ctf_support(filepath: str | Path) -> dict[str, Any]:
    """
    Parse a CTF-FTC-Support Excel file.

    Returns dict with keys:
      technologySupport, capacityBuilding, supportProjects (merged/deduplicated)
    """
    filepath = Path(filepath)
    logger.info(f"Parsing CTF-FTC-Support: {filepath.name}")
    wb = load_workbook(filepath, data_only=True, read_only=True)

    tech: list[dict[str, Any]] = []
    cb: list[dict[str, Any]] = []

    # Table 9 in FTC = technology support received
    ws9 = _find_sheet(wb, ["table9"])
    if ws9:
        tech = _parse_support_table(ws9)
        logger.info(f"  Table9: {len(tech)} technology support projects")

    # Table 11 in FTC = capacity building received
    ws11 = _find_sheet(wb, ["table11"])
    if ws11:
        cb = _parse_support_table(ws11)
        logger.info(f"  Table11: {len(cb)} capacity-building projects")

    merged = _merge_support_projects(tech, cb)
    logger.info(f"  Merged: {len(merged)} unique support projects")

    wb.close()
    return {
        "sourceFile": filepath.name,
        "technologySupport": tech,
        "capacityBuilding": cb,
        "supportProjects": merged,
    }


# ---------------------------------------------------------------------------
# CLI entry point for testing
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import sys
    logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")

    if len(sys.argv) < 2:
        print("Usage: python -m src.parse_ctf <ctf-excel-file> [--type ndc|support]")
        sys.exit(1)

    fpath = Path(sys.argv[1])
    ftype = "ndc"
    if "--type" in sys.argv:
        idx = sys.argv.index("--type")
        ftype = sys.argv[idx + 1] if idx + 1 < len(sys.argv) else "ndc"

    if ftype == "support":
        data = parse_ctf_support(fpath)
    elif ftype == "all":
        data = parse_ctf_ndc(fpath)
        support = parse_ctf_support(fpath)
        data["technologySupport"] = support.get("technologySupport", [])
        data["capacityBuilding"] = support.get("capacityBuilding", [])
    else:
        data = parse_ctf_ndc(fpath)

    print(json.dumps(data, indent=2, default=str))
