"""
Parse a general target-list Excel file into target rows.

Auto-detects column headers (Target Name, Target Text, Activities,
Measures/Action, Target Type, etc.) and returns structured target rows
including optional activities and actions fields.

Usage:
    python -m src.parse_targets_excel --file path/to/targets.xlsx [--sheet "Sheet1"]
"""

from __future__ import annotations

import argparse
import json
import logging
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

# Known document type mappings (lowercase prefix → canonical type)
KNOWN_DOC_TYPES: dict[str, str] = {
    "ndc": "NDC",
    "nbsap": "NBSAP",
    "nbt": "NBSAP",
    "national biodiversity": "NBSAP",
    "nap": "NAP",
    "ldn": "LDN",
    "sectoral": "SECTORAL",
    "other": "OTHER",
}


def _match_doc_type(raw: str) -> str:
    lower = raw.lower().strip()
    for keyword, doc_type in KNOWN_DOC_TYPES.items():
        if lower == keyword or lower.startswith(keyword):
            return doc_type
    return "OTHER"


def _detect_columns(headers: list[str]) -> dict[str, int]:
    """Detect column roles from header strings. Returns col name → index."""
    mapping: dict[str, int] = {}
    lower = [h.lower().strip() for h in headers]

    for i, h in enumerate(lower):
        if not h:
            continue
        if "target text" in h and "aggregate" not in h:
            mapping["text"] = i
        elif h in ("text", "description"):
            mapping.setdefault("text", i)
        elif "target type" in h or "document type" in h or "doc type" in h or h == "type" or h == "source document":
            mapping["docType"] = i
        elif "target name" in h or h in ("label", "name", "target"):
            mapping["label"] = i
        elif "activities" in h or "activity" in h:
            mapping["activities"] = i
        elif "measures" in h or ("action" in h and "target text" not in h):
            mapping["actions"] = i

    return mapping


def parse_targets_excel(
    file_path: str | Path,
    sheet_name: str | None = None,
) -> list[dict[str, Any]]:
    """Parse an Excel file into a list of target dicts."""
    wb = load_workbook(file_path, read_only=True, data_only=True)

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
        ws = wb[sheet_name]
    else:
        ws = wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return []

    headers = [str(h) if h else "" for h in rows[0]]
    data = rows[1:]

    mapping = _detect_columns(headers)

    if "text" not in mapping:
        # Fallback: pick the column with the longest average content
        best_col = 0
        best_avg = 0.0
        for c in range(len(headers)):
            avg = sum(len(str(r[c] or "")) for r in data[:10]) / max(len(data[:10]), 1)
            if avg > best_avg:
                best_avg = avg
                best_col = c
        mapping["text"] = best_col

    targets = []
    for row in data:
        text_idx = mapping["text"]
        text = str(row[text_idx] or "").strip() if text_idx < len(row) else ""
        if not text:
            continue

        label = ""
        if "label" in mapping and mapping["label"] < len(row):
            label = str(row[mapping["label"]] or "").strip()

        doc_type = "OTHER"
        if "docType" in mapping and mapping["docType"] < len(row):
            raw_type = str(row[mapping["docType"]] or "").strip()
            if raw_type:
                doc_type = _match_doc_type(raw_type)

        target: dict[str, Any] = {
            "text": text,
            "sourceDocument": doc_type,
            "sourceLabel": label or f"Target {len(targets) + 1}",
        }

        if "activities" in mapping and mapping["activities"] < len(row):
            val = str(row[mapping["activities"]] or "").strip()
            if val:
                target["activities"] = val

        if "actions" in mapping and mapping["actions"] < len(row):
            val = str(row[mapping["actions"]] or "").strip()
            if val:
                target["actions"] = val

        targets.append(target)

    return targets


def main():
    parser = argparse.ArgumentParser(description="Parse target-list Excel into JSON")
    parser.add_argument("--file", required=True, help="Path to Excel file")
    parser.add_argument("--sheet", default=None, help="Sheet name (default: first sheet)")
    parser.add_argument("--output", default=None, help="Output JSON path (default: stdout)")
    args = parser.parse_args()

    targets = parse_targets_excel(args.file, args.sheet)

    out = json.dumps(targets, indent=2, ensure_ascii=False)
    if args.output:
        Path(args.output).write_text(out)
        print(f"Wrote {len(targets)} targets to {args.output}", file=sys.stderr)
    else:
        print(out)


if __name__ == "__main__":
    main()
