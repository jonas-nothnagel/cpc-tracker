"""
Build a comprehensive XLSX report of Mongolia BER data, GLOBE taxonomy,
expert calibration examples, and LLM-generated classifications.

Output: python/output/mongolia/Mongolia_BER_GLOBE_Report.xlsx

Usage:
    cd python
    python scripts/build_mongolia_ber_report.py
"""

from __future__ import annotations

import json
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

PYTHON_DIR = Path(__file__).resolve().parent.parent
DATA_DIR = PYTHON_DIR / "data"
OUT_DIR = PYTHON_DIR / "output" / "mongolia"
OUT_DIR.mkdir(parents=True, exist_ok=True)
OUT_FILE = OUT_DIR / "Mongolia_BER_GLOBE_Report.xlsx"

# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="1E40AF")  # UNDP blue
HEADER_FONT = Font(color="FFFFFF", bold=True)
PARENT_FILL = PatternFill("solid", fgColor="DBEAFE")  # light blue
SECTION_FONT = Font(bold=True, size=11)
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")
TOP_ALIGN = Alignment(vertical="top")
THIN_BORDER = Border(
    left=Side(style="thin", color="E5E7EB"),
    right=Side(style="thin", color="E5E7EB"),
    top=Side(style="thin", color="E5E7EB"),
    bottom=Side(style="thin", color="E5E7EB"),
)


def style_header_row(ws: Worksheet, row: int, num_cols: int) -> None:
    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[row].height = 28


def set_col_widths(ws: Worksheet, widths: list[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------


def load_data() -> dict:
    categories = json.loads((DATA_DIR / "categories.json").read_text())
    ber = json.loads((DATA_DIR / "mongolia-ber.json").read_text())
    ber_tax = json.loads((DATA_DIR / "mongolia-ber-taxonomy.json").read_text())
    ber_purpose = json.loads((DATA_DIR / "mongolia-ber-purpose-allocations.json").read_text())
    examples = json.loads((DATA_DIR / "mongolia-ber-globe-examples.json").read_text())
    targets = json.loads((DATA_DIR / "mongolia-targets.json").read_text())
    classifications = json.loads((OUT_DIR / "classifications.json").read_text())
    budget_pseudo = json.loads((OUT_DIR / "budget_pseudo_targets.json").read_text())
    measure_pseudo = json.loads((OUT_DIR / "measure_pseudo_targets.json").read_text())

    return {
        "globe_categories": categories["globe_categories"],
        "globe_subcategories": categories["globe_subcategories"],
        "ber": ber,
        "ber_taxonomy": ber_tax,
        "ber_purpose": ber_purpose,
        "examples": examples,
        "targets": targets,
        "classifications": classifications,
        "budget_pseudo": budget_pseudo,
        "measure_pseudo": measure_pseudo,
    }


# ---------------------------------------------------------------------------
# Sheet: README
# ---------------------------------------------------------------------------


def build_readme(wb: Workbook, d: dict) -> None:
    ws = wb.create_sheet("README")

    rows = [
        ("Mongolia BER × BIOFIN GLOBE Taxonomy Report", ""),
        ("", ""),
        ("Purpose", "Shared reference combining Mongolia's Biodiversity Expenditure Review (BER), the BIOFIN/UNDP GLOBE 2024 taxonomy, expert-curated classification examples, and LLM-generated subcategory classifications across policy targets, BTR actions, and budget programs."),
        ("", ""),
        ("Source documents", "• BER report: BER [ENG] v1.2.docx (Unet Consulting LLC, Nov 2025)\n• BER spreadsheet: BER_1112.xlsx (249 MB)\n• GLOBE taxonomy: https://www.biofin.org/knowledge-product/global-biodiversity-expenditure-taxonomy\n• Pipeline: github.com/undp/cpc-tracker, branch feat/ber-financing-integration"),
        ("", ""),
        ("Sheet guide", ""),
        ("1. GLOBE Taxonomy", "The full BIOFIN GLOBE 2024 taxonomy — 9 Primary Biodiversity Categories and 48 subcategories, with descriptions drawn from the GLOBE 2024 guidance document."),
        ("2. Public Sector BER", "Mongolia's 28 budgetary biodiversity programs (from the BER Word doc). Includes English name, description, type flag, yearly expenditure 2020--2024, and LLM-generated GLOBE subcategory assignments."),
        ("2c. Biodiversity Expenditures", "The 15 biodiversity-specific public expenditure sub-codes (e.g. 81707 air pollution, 81708 afforestation, 81703 plant protection). These are the core lines the researcher identified as biodiversity-attributable in the BER analysis, with full 2016-2024 yearly detail and detailed accounting rules. Mongolian original translated to English."),
        ("2b. BAR Sub-categories", "58 economic classification sub-codes (e.g. 80101, 81707) from the BAR sheet of the BER Excel, translated from Mongolian to English. These are the sub-codes the researcher used to attribute each line of public expenditure with a biodiversity relevance rate (100% / 75% / 50% / 5%). Sub-codes are a parallel classification (HOW money is spent) to the program codes (WHAT area)."),
        ("3. BER Classifications", "Detailed LLM classifications of each BER program against GLOBE subcategories. Shows only the matched subcategories with confidence level and reasoning."),
        ("4. Target Classifications", "LLM classifications of the 77 policy targets (from NBSAP, NDC, NAP, etc.) against GLOBE subcategories -- same cross-level taxonomy bridge."),
        ("5. Action Classifications", "LLM classifications of the 39 implementation actions (BTR mitigation + APNDC adaptation) against GLOBE subcategories."),
        ("6. Expert Examples (Private)", "The 138 expert-curated GLOBE mappings from the BER Excel Sheet 5. These are the Top 100 Mongolian companies' biodiversity expenditures classified by BIOFIN experts. Descriptions are shown in both English (Claude translation) and the original Mongolian. These are the few-shot calibration examples used by the LLM classifier."),
        ("7. Cross-level Spine", "For each GLOBE subcategory: which budget programs fund it, which policy targets address it, and which BTR/ADP actions implement it. The central cross-level coherence view."),
        ("8. Summary Stats", "Coverage statistics, confidence distributions, top-used subcategories, and headline financial figures."),
        ("", ""),
        ("Methodology note (LLM classification)", "The LLM classifier (calibrated on the 138 expert examples in Sheet 6) assigns each item to one or more GLOBE subcategories with a confidence level (High/Medium/Low) and short reasoning. 'Low' confidence is normal and matches the expert pattern — most private sector items received Low confidence because biodiversity intent is often indirect. Public sector programs tend toward Medium/High because they have more explicit biodiversity objectives."),
        ("", ""),
        ("Open question for the team", "The BER report says the researcher 'operationalized using the BIOFIN/GLOBE taxonomy' for the public sector but the subcategory codes were NOT recorded in the Excel. Does a separate structured GLOBE mapping exist for the 28 public sector programs? If yes, we can replace/audit the LLM classifications against ground truth. If no, this report documents a path to automate that tagging."),
    ]

    for i, (key, value) in enumerate(rows, start=1):
        ws.cell(row=i, column=1, value=key).font = Font(bold=True, size=11 if i == 1 else 10)
        if i == 1:
            ws.cell(row=i, column=1).font = Font(bold=True, size=14, color="1E40AF")
        if value:
            c = ws.cell(row=i, column=2, value=value)
            c.alignment = WRAP_ALIGN

    set_col_widths(ws, [32, 95])
    # Set row heights for long paragraphs
    for i in range(1, len(rows) + 1):
        ws.row_dimensions[i].height = max(18, 15 * (len(rows[i - 1][1]) // 70 + 1))


# ---------------------------------------------------------------------------
# Sheet 1: GLOBE taxonomy
# ---------------------------------------------------------------------------


def build_taxonomy_sheet(wb: Workbook, d: dict) -> None:
    ws = wb.create_sheet("1. GLOBE Taxonomy")
    headers = [
        "Code",
        "Primary Category",
        "Subcategory",
        "Description",
    ]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))

    parent_by_id = {c["id"]: c for c in d["globe_categories"]}

    # Section-by-section for readability
    grouped: dict[str, list[dict]] = defaultdict(list)
    for sub in d["globe_subcategories"]:
        grouped[sub["parentId"]].append(sub)

    for parent_id in sorted(grouped.keys()):
        parent = parent_by_id[parent_id]
        # Primary category header row
        row = [
            parent_id.replace("globe_", ""),
            parent["name"].upper(),
            "",
            parent["description"],
        ]
        ws.append(row)
        for col_idx in range(1, 5):
            cell = ws.cell(row=ws.max_row, column=col_idx)
            cell.fill = PARENT_FILL
            cell.font = SECTION_FONT
            cell.alignment = WRAP_ALIGN

        for sub in sorted(grouped[parent_id], key=lambda s: s["id"]):
            ws.append([sub["id"], parent["name"], sub["name"], sub["description"]])
            for col_idx in range(1, 5):
                ws.cell(row=ws.max_row, column=col_idx).alignment = WRAP_ALIGN

    set_col_widths(ws, [8, 30, 42, 90])
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Sheet 2: Public Sector BER programs
# ---------------------------------------------------------------------------


def build_ber_programs_sheet(wb: Workbook, d: dict) -> None:
    ws = wb.create_sheet("2. Public Sector BER")

    years = sorted({y for e in d["ber"]["expenditure"] for y in e["values"].keys()})
    headers = [
        "Program Code",
        "Name (EN)",
        "Type",
        "Has Expenditure",
        "Description",
        *[f"{y} ({d['ber']['unit']} {d['ber']['currency']})" for y in years],
        f"Total {years[0]}-{years[-1]}",
        "Assigned GLOBE Subcategories (LLM)",
    ]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))

    # Build expenditure lookup
    exp_by_code = {e["code"]: e for e in d["ber"]["expenditure"]}

    # Build GLOBE assignment lookup per BER program
    ber_globe: dict[str, list[tuple[str, str, str]]] = defaultdict(list)
    for cls in d["classifications"]:
        if cls.get("taxonomyType") != "globe_sub":
            continue
        if not cls.get("isRelevant"):
            continue
        if not cls["targetId"].startswith("BER_"):
            continue
        code = cls["targetId"].removeprefix("BER_")
        ber_globe[code].append(
            (cls["categoryId"], cls.get("confidence", ""), cls.get("reasoning", ""))
        )

    for prog in d["ber"]["programs"]:
        code = prog["code"]
        exp = exp_by_code.get(code, {"values": {}})
        values = exp["values"]
        total = sum(v for v in values.values() if v is not None)
        has_exp = any(v is not None and v > 0 for v in values.values())
        year_vals = [values.get(y, None) for y in years]

        globe_assignments = ber_globe.get(code, [])
        globe_str = (
            "; ".join(f"{sub} [{conf}]" for sub, conf, _ in globe_assignments)
            if globe_assignments
            else ""
        )

        row = [
            code,
            prog.get("name", ""),
            prog.get("type", ""),
            "Yes" if has_exp else "No",
            prog.get("description", ""),
            *year_vals,
            round(total, 3) if total else 0,
            globe_str,
        ]
        ws.append(row)
        # Wrap description and globe columns
        r = ws.max_row
        for col_idx in [2, 5, len(headers)]:
            ws.cell(row=r, column=col_idx).alignment = WRAP_ALIGN

    widths = [12, 32, 16, 14, 50, *[12] * len(years), 16, 45]
    set_col_widths(ws, widths)
    ws.freeze_panes = "B2"


# ---------------------------------------------------------------------------
# Sheet 3: BER GLOBE classifications (long-form)
# ---------------------------------------------------------------------------


def build_ber_classifications_sheet(wb: Workbook, d: dict) -> None:
    ws = wb.create_sheet("3. BER Classifications")
    headers = [
        "Program Code",
        "Program Name",
        "Subcategory Code",
        "Subcategory Name",
        "Primary Category",
        "Confidence",
        "Reasoning",
    ]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))

    program_by_code = {p["code"]: p for p in d["ber"]["programs"]}
    sub_by_id = {s["id"]: s for s in d["globe_subcategories"]}
    parent_by_id = {c["id"]: c["name"] for c in d["globe_categories"]}

    rows_added = 0
    for cls in d["classifications"]:
        if cls.get("taxonomyType") != "globe_sub":
            continue
        if not cls.get("isRelevant"):
            continue
        if not cls["targetId"].startswith("BER_"):
            continue

        code = cls["targetId"].removeprefix("BER_")
        prog = program_by_code.get(code, {})
        sub = sub_by_id.get(cls["categoryId"], {})
        parent_name = parent_by_id.get(sub.get("parentId", ""), "")

        ws.append([
            code,
            prog.get("name", ""),
            cls["categoryId"],
            sub.get("name", ""),
            parent_name,
            cls.get("confidence", ""),
            cls.get("reasoning", ""),
        ])
        for col_idx in [2, 4, 5, 7]:
            ws.cell(row=ws.max_row, column=col_idx).alignment = WRAP_ALIGN
        rows_added += 1

    set_col_widths(ws, [12, 32, 14, 35, 28, 12, 60])
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Sheet 4: Policy target classifications
# ---------------------------------------------------------------------------


def build_target_classifications_sheet(wb: Workbook, d: dict) -> None:
    ws = wb.create_sheet("4. Target Classifications")
    headers = [
        "Target ID",
        "Source Document",
        "Target Text",
        "Subcategory Code",
        "Subcategory Name",
        "Primary Category",
        "Confidence",
        "Reasoning",
    ]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))

    target_by_id = {t["id"]: t for t in d["targets"]}
    sub_by_id = {s["id"]: s for s in d["globe_subcategories"]}
    parent_by_id = {c["id"]: c["name"] for c in d["globe_categories"]}

    for cls in d["classifications"]:
        if cls.get("taxonomyType") != "globe_sub":
            continue
        if not cls.get("isRelevant"):
            continue
        tid = cls["targetId"]
        if tid.startswith("BER_") or tid.startswith("BTR_") or tid.startswith("ADP_"):
            continue
        target = target_by_id.get(tid)
        if not target:
            continue
        sub = sub_by_id.get(cls["categoryId"], {})
        parent_name = parent_by_id.get(sub.get("parentId", ""), "")

        ws.append([
            tid,
            target.get("sourceDocument", ""),
            target.get("text", ""),
            cls["categoryId"],
            sub.get("name", ""),
            parent_name,
            cls.get("confidence", ""),
            cls.get("reasoning", ""),
        ])
        for col_idx in [3, 5, 6, 8]:
            ws.cell(row=ws.max_row, column=col_idx).alignment = WRAP_ALIGN

    set_col_widths(ws, [14, 16, 60, 14, 35, 28, 12, 55])
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Sheet 5: BTR + ADP action classifications
# ---------------------------------------------------------------------------


def build_action_classifications_sheet(wb: Workbook, d: dict) -> None:
    ws = wb.create_sheet("5. Action Classifications")
    headers = [
        "Action ID",
        "Action Type",
        "Source Label",
        "Description",
        "Subcategory Code",
        "Subcategory Name",
        "Primary Category",
        "Confidence",
        "Reasoning",
    ]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))

    measure_by_id = {m["id"]: m for m in d["measure_pseudo"]}
    sub_by_id = {s["id"]: s for s in d["globe_subcategories"]}
    parent_by_id = {c["id"]: c["name"] for c in d["globe_categories"]}

    for cls in d["classifications"]:
        if cls.get("taxonomyType") != "globe_sub":
            continue
        if not cls.get("isRelevant"):
            continue
        tid = cls["targetId"]
        if not (tid.startswith("BTR_") or tid.startswith("ADP_")):
            continue
        action = measure_by_id.get(tid)
        if not action:
            continue
        sub = sub_by_id.get(cls["categoryId"], {})
        parent_name = parent_by_id.get(sub.get("parentId", ""), "")

        ws.append([
            tid,
            action.get("actionType", ""),
            action.get("sourceLabel", ""),
            action.get("text", ""),
            cls["categoryId"],
            sub.get("name", ""),
            parent_name,
            cls.get("confidence", ""),
            cls.get("reasoning", ""),
        ])
        for col_idx in [3, 4, 6, 7, 9]:
            ws.cell(row=ws.max_row, column=col_idx).alignment = WRAP_ALIGN

    set_col_widths(ws, [12, 14, 35, 55, 14, 30, 28, 12, 50])
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Sheet 6: Expert examples (private sector)
# ---------------------------------------------------------------------------


def build_expert_examples_sheet(wb: Workbook, d: dict) -> None:
    ws = wb.create_sheet("6. Expert Examples (Private)")
    headers = [
        "Row ID",
        "Description (English)",
        "Description (Mongolian)",
        "GLOBE Subcategories",
        "Intent Label",
        "BAR Justification",
        "Confidence",
        "Start Year",
        "End Year",
        "Expenditure Before BAR (MNT)",
        "BAR Rate",
        "Expenditure After BAR (MNT)",
    ]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))

    for rec in d["examples"].get("records", []):
        codes = ", ".join(
            f"{c['code']} {c.get('name', '')}"
            for c in rec.get("globeSubcategories", [])
        )
        ws.append([
            rec.get("id", ""),
            rec.get("descriptionEn", ""),
            rec.get("description", ""),
            codes,
            rec.get("intentLabel", ""),
            rec.get("barJustification", ""),
            rec.get("confidence", ""),
            rec.get("startYear", ""),
            rec.get("endYear", ""),
            rec.get("expenditureBeforeBarMNT"),
            rec.get("barRate"),
            rec.get("expenditureAfterBarMNT"),
        ])
        for col_idx in [2, 3, 4, 5, 6]:
            ws.cell(row=ws.max_row, column=col_idx).alignment = WRAP_ALIGN

    set_col_widths(ws, [8, 55, 55, 40, 35, 40, 12, 10, 10, 22, 10, 22])
    ws.freeze_panes = "A2"


def build_purpose_allocations_sheet(wb: Workbook, d: dict) -> None:
    """Sheet 2c: The 15 biodiversity-specific expenditure items from Sheet 2 of
    the BER Excel (Зориулалт). These are the core public-sector biodiversity
    expenditure lines the researcher used to calculate BAR-adjusted totals,
    with full 2016-2024 yearly breakdown. Mongolian translated to English.
    """
    ws = wb.create_sheet("2c. Biodiversity Expenditures")

    purpose = d["ber_purpose"]
    items = purpose.get("items", [])
    years = sorted(
        {y for item in items for y in item.get("expenditureByYearBillionMnt", {}).keys()}
    )

    headers = [
        "Sub-code",
        "Name (English)",
        "Name (Mongolian)",
        "Description (English)",
        "Description (Mongolian)",
        "BAR Rate",
        *[f"{y} (B MNT)" for y in years],
        "Total (B MNT)",
    ]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))

    for item in items:
        yearly = item.get("expenditureByYearBillionMnt", {})
        year_vals = [yearly.get(y) for y in years]
        row = [
            item.get("code", ""),
            item.get("nameEn", ""),
            item.get("nameMn", ""),
            item.get("descriptionEn", ""),
            item.get("descriptionMn", ""),
            item.get("barRate"),
            *year_vals,
            item.get("totalBillionMnt"),
        ]
        ws.append(row)
        r = ws.max_row
        for col_idx in [2, 3, 4, 5]:
            ws.cell(row=r, column=col_idx).alignment = WRAP_ALIGN

    set_col_widths(ws, [10, 42, 42, 60, 60, 10, *[10] * len(years), 14])
    ws.freeze_panes = "B2"


def build_bar_translations_sheet(wb: Workbook, d: dict) -> None:
    """Sheet with the 58 BAR sub-category translations from mongolia-ber-taxonomy.json.

    These are economic classification sub-codes (e.g. 80101, 81707) from the BAR
    sheet of the BER Excel — the sub-categories the researcher used to attribute
    each line of expenditure with a biodiversity relevance rate (100% / 75% / 50% /
    5%). They sit alongside (not inside) the 28 program codes: programs say WHAT
    area, BAR sub-codes say HOW money is spent.
    """
    ws = wb.create_sheet("2b. BAR Sub-categories")

    headers = [
        "Sub-code",
        "Description (English)",
        "Description (Mongolian)",
        "BIOFIN Relevance",
        "Total Expenditure 2020-2024 (B MNT)",
    ]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))

    taxonomy = d["ber_taxonomy"]
    sub_categories = taxonomy.get("subCategories", [])

    # Sort by relevance desc, then by code
    sorted_subs = sorted(
        sub_categories,
        key=lambda s: (-(s.get("biofinRelevance") or 0), s.get("code", "")),
    )

    # Group header rows by relevance tier for readability
    current_tier: float | None = None
    tier_labels = {
        1.0: "100% BIOFIN relevance -- fully biodiversity-attributable",
        0.75: "75% BIOFIN relevance -- predominantly biodiversity-related",
        0.5: "50% BIOFIN relevance -- partial biodiversity relevance",
        0.05: "5% BIOFIN relevance -- tangential environmental link",
        0.0: "0% BIOFIN relevance",
    }

    for sc in sorted_subs:
        rel = sc.get("biofinRelevance")
        if rel != current_tier:
            current_tier = rel
            label = tier_labels.get(rel, f"BIOFIN relevance: {rel}")
            ws.append([label, "", "", "", ""])
            r = ws.max_row
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=r, column=col_idx)
                cell.fill = PARENT_FILL
                cell.font = SECTION_FONT
                cell.alignment = WRAP_ALIGN

        total = sc.get("totalExpenditureBillionMnt")
        ws.append([
            sc.get("code", ""),
            sc.get("descriptionEn", ""),
            sc.get("descriptionMn", ""),
            rel,
            total,
        ])
        r = ws.max_row
        for col_idx in [2, 3]:
            ws.cell(row=r, column=col_idx).alignment = WRAP_ALIGN

    set_col_widths(ws, [12, 55, 55, 18, 22])
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Sheet 7: Cross-level spine
# ---------------------------------------------------------------------------


def build_cross_level_sheet(wb: Workbook, d: dict) -> None:
    ws = wb.create_sheet("7. Cross-level Spine")
    headers = [
        "Subcategory Code",
        "Subcategory Name",
        "Primary Category",
        "# BER Programs",
        "BER Program Codes",
        "Total Expenditure 2020-2024 (B MNT)",
        "# Policy Targets",
        "Target IDs",
        "# BTR+ADP Actions",
        "Action IDs",
    ]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))

    parent_by_id = {c["id"]: c["name"] for c in d["globe_categories"]}
    sub_by_id = {s["id"]: s for s in d["globe_subcategories"]}

    # Expenditure per BER code (cumulative 2020-2024)
    exp_by_code: dict[str, float] = {}
    for e in d["ber"]["expenditure"]:
        exp_by_code[e["code"]] = sum(v for v in e["values"].values() if v is not None)

    # Build per-subcategory rollup
    per_sub: dict[str, dict] = defaultdict(
        lambda: {"ber": [], "targets": [], "actions": []}
    )
    for cls in d["classifications"]:
        if cls.get("taxonomyType") != "globe_sub":
            continue
        if not cls.get("isRelevant"):
            continue
        sub_id = cls["categoryId"]
        tid = cls["targetId"]
        if tid.startswith("BER_"):
            per_sub[sub_id]["ber"].append(tid.removeprefix("BER_"))
        elif tid.startswith("BTR_") or tid.startswith("ADP_"):
            per_sub[sub_id]["actions"].append(tid)
        else:
            per_sub[sub_id]["targets"].append(tid)

    # Emit one row per subcategory in canonical order
    for sub in sorted(d["globe_subcategories"], key=lambda s: s["id"]):
        sid = sub["id"]
        data = per_sub.get(sid, {"ber": [], "targets": [], "actions": []})
        total_exp = sum(exp_by_code.get(code, 0) for code in data["ber"])
        row = [
            sid,
            sub["name"],
            parent_by_id.get(sub["parentId"], ""),
            len(data["ber"]),
            ", ".join(sorted(set(data["ber"]))),
            round(total_exp, 3) if total_exp else 0,
            len(set(data["targets"])),
            ", ".join(sorted(set(data["targets"]))),
            len(set(data["actions"])),
            ", ".join(sorted(set(data["actions"]))),
        ]
        ws.append(row)
        for col_idx in [2, 3, 5, 8, 10]:
            ws.cell(row=ws.max_row, column=col_idx).alignment = WRAP_ALIGN

    set_col_widths(ws, [14, 35, 28, 14, 30, 16, 14, 50, 14, 50])
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Sheet 8: Summary statistics
# ---------------------------------------------------------------------------


def build_summary_sheet(wb: Workbook, d: dict) -> None:
    ws = wb.create_sheet("8. Summary Stats")

    # Sub classifications
    sub_cls = [
        c
        for c in d["classifications"]
        if c.get("taxonomyType") == "globe_sub" and c.get("isRelevant")
    ]

    ber_cls = [c for c in sub_cls if c["targetId"].startswith("BER_")]
    action_cls = [c for c in sub_cls if c["targetId"].startswith(("BTR_", "ADP_"))]
    target_cls = [
        c for c in sub_cls if not c["targetId"].startswith(("BER_", "BTR_", "ADP_"))
    ]

    # Expenditure totals
    total_exp_all = 0.0
    total_exp_classified = 0.0
    for e in d["ber"]["expenditure"]:
        exp_sum = sum(v for v in e["values"].values() if v is not None)
        total_exp_all += exp_sum
        code = e["code"]
        if any(c["targetId"] == f"BER_{code}" for c in ber_cls):
            total_exp_classified += exp_sum

    # Confidence distribution
    conf_counts: dict[str, int] = defaultdict(int)
    for c in sub_cls:
        conf_counts[c.get("confidence", "unspecified")] += 1

    # Top subcategories by usage
    sub_usage: dict[str, int] = defaultdict(int)
    for c in sub_cls:
        sub_usage[c["categoryId"]] += 1
    top_subs = sorted(sub_usage.items(), key=lambda x: -x[1])[:15]
    sub_by_id = {s["id"]: s for s in d["globe_subcategories"]}

    rows = [
        ("Section", "Metric", "Value"),
        ("Input data", "Number of BER programs (public sector)", len(d["ber"]["programs"])),
        ("Input data", "Programs with 2020-2024 expenditure", sum(1 for e in d["ber"]["expenditure"] if any(v for v in e["values"].values() if v))),
        ("Input data", "Policy targets", len(d["targets"])),
        ("Input data", "BTR + ADP implementation actions", len(d["measure_pseudo"])),
        ("Input data", "Expert example classifications (private sector)", len(d["examples"].get("records", []))),
        ("Input data", "GLOBE primary categories", len(d["globe_categories"])),
        ("Input data", "GLOBE subcategories", len(d["globe_subcategories"])),
        ("", "", ""),
        ("Financial coverage", f"Total public BER expenditure 2020-2024 ({d['ber']['unit']} {d['ber']['currency']})", round(total_exp_all, 1)),
        ("Financial coverage", f"Of which classified to GLOBE subcategories", round(total_exp_classified, 1)),
        ("Financial coverage", "Coverage %", f"{(100 * total_exp_classified / total_exp_all if total_exp_all else 0):.1f}%"),
        ("", "", ""),
        ("Classification output", "Total relevant subcategory matches", len(sub_cls)),
        ("Classification output", "— from BER programs", len(ber_cls)),
        ("Classification output", "— from policy targets", len(target_cls)),
        ("Classification output", "— from BTR + ADP actions", len(action_cls)),
        ("Classification output", "Avg subcategories per BER program", round(len(ber_cls) / max(1, sum(1 for e in d["ber"]["expenditure"] if any(v for v in e["values"].values() if v))), 2)),
        ("", "", ""),
        ("Confidence distribution", "High", conf_counts.get("High", 0)),
        ("Confidence distribution", "Medium", conf_counts.get("Medium", 0)),
        ("Confidence distribution", "Low", conf_counts.get("Low", 0)),
        ("", "", ""),
        ("Top 15 subcategories used (overall)", "", ""),
    ]
    for code, count in top_subs:
        name = sub_by_id.get(code, {}).get("name", "")
        rows.append((code, name, count))

    headers = rows[0]
    ws.append(list(headers))
    style_header_row(ws, 1, len(headers))

    for r in rows[1:]:
        ws.append(list(r))
        is_section = r[0] and r[1] == "" and r[2] == ""
        if is_section:
            for col in range(1, 4):
                ws.cell(row=ws.max_row, column=col).font = Font(bold=True, size=11, color="1E40AF")

    set_col_widths(ws, [32, 45, 22])
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main() -> None:
    d = load_data()

    wb = Workbook()
    # Remove default empty sheet
    default = wb.active
    wb.remove(default)

    build_readme(wb, d)
    build_taxonomy_sheet(wb, d)
    build_ber_programs_sheet(wb, d)
    build_purpose_allocations_sheet(wb, d)
    build_bar_translations_sheet(wb, d)
    build_ber_classifications_sheet(wb, d)
    build_target_classifications_sheet(wb, d)
    build_action_classifications_sheet(wb, d)
    build_expert_examples_sheet(wb, d)
    build_cross_level_sheet(wb, d)
    build_summary_sheet(wb, d)

    wb.save(OUT_FILE)
    print(f"Wrote {OUT_FILE}")
    print(f"Size: {OUT_FILE.stat().st_size / 1024:.1f} KB")
    for sheet in wb.sheetnames:
        print(f"  - {sheet}: {wb[sheet].max_row} rows")


if __name__ == "__main__":
    main()
