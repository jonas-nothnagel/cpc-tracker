#!/usr/bin/env python3
"""One-off ingestion of Mongolia LDN (land-degradation) targets.

Source: dev_data_scripts/sharepoint_sync/Mongolia/Mongolia materials/
        LDN targets/land_targets.xlsx  (sheet "targets")

The sheet holds two UNCCD land-degradation documents, one target per row,
grouped by column F (Doc code) / column D (Document title):

  NRVTS  National Report on Voluntary Target Setting to Achieve Land
         Degradation Neutrality in Mongolia            -> 27 rows
  ILDN   Investing in Land Degradation Neutrality      ->  8 rows

Each row becomes one target matching the schema in
python/data/mongolia-targets.json:
  {id, text, sourceDocument, sourceLabel, country, sources[], textCleanup}

Text fidelity: the verbatim row text is preserved in sources[].sourceText.
A small, explicit list of missing-space PDF-extraction fixes is applied to
the display `text` only (no wording changes); fixed rows are flagged
textCleanup="cleaned", untouched rows "verbatim".

This is a one-off; the generic parse_targets_excel.py does not handle this
sheet's headers / multi-document layout. The document-upload pipeline is a
separate effort.

Usage (from python/):
    python scripts/ingest_ldn_targets.py            # write preview + QC report
    python scripts/ingest_ldn_targets.py --append   # append into mongolia-targets.json
"""
from __future__ import annotations

import argparse
import json
import re
from pathlib import Path

from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parents[2]
XLSX = (
    REPO_ROOT
    / "dev_data_scripts/sharepoint_sync/Mongolia/Mongolia materials/LDN targets/land_targets.xlsx"
)
DATA_DIR = Path(__file__).resolve().parents[1] / "data"
TARGETS_FILE = DATA_DIR / "mongolia-targets.json"
PREVIEW_FILE = DATA_DIR / "_ldn_targets_preview.json"

# Human-readable provenance for the country config docProvenance (printed for
# convenience; the config edit itself is done by hand to keep it reviewable).
SOURCE_PDF = "dev_data_scripts/sharepoint_sync/Mongolia/Mongolia materials/LDN targets/Mongolia LDN TSP Country Report.pdf"

# Explicit, auditable missing-space repairs. Each key is a unique substring in
# the affected NRVTS rows; nothing here matches a clean target. Applied to the
# display `text` only; the original stays in sources[].sourceText.
SPACING_FIXES: dict[str, str] = {
    "inareas": "in areas",
    "systemfor": "system for",
    "foreststeppe": "forest steppe",
    "perannum": "per annum",
    "1.6t/ha": "1.6 t/ha",
    "annumin": "annum in",
    "ofpesticides": "of pesticides",
    "suitablesystem": "suitable system",
    "forecosystem": "for ecosystem",
    "inheadwaters": "in headwaters",
    "andwetlands": "and wetlands",
}


def _normalize_ws(s: str) -> str:
    """Collapse newlines / runs of whitespace to single spaces, strip ends."""
    return re.sub(r"\s+", " ", s).strip()


def _apply_fixes(text: str) -> tuple[str, list[tuple[str, str]]]:
    """Apply the explicit spacing fixes. Returns (fixed_text, applied)."""
    applied: list[tuple[str, str]] = []
    out = text
    for bad, good in SPACING_FIXES.items():
        if bad in out:
            out = out.replace(bad, good)
            applied.append((bad, good))
    return out, applied


def parse() -> list[dict]:
    wb = load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb["targets"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    targets: list[dict] = []
    counters: dict[str, int] = {}
    for r in rows[1:]:  # skip header
        title, text, document, _conv, code = r[1], r[2], r[3], r[4], r[5]
        if not any([title, text, document]):
            continue  # trailing empty rows
        code = str(code).strip()
        verbatim = str(text).strip()
        fixed, applied = _apply_fixes(verbatim)

        counters[code] = counters.get(code, 0) + 1
        targets.append(
            {
                "id": f"{code}_{counters[code]}",
                "text": fixed,
                "sourceDocument": code,
                "sourceLabel": _normalize_ws(str(title)),
                "country": "Mongolia",
                "sources": [{"sourceText": verbatim}],
                "textCleanup": "cleaned" if applied else "verbatim",
                # transient QC field, stripped before append:
                "_fixes": applied,
            }
        )
    return targets


def qc_report(targets: list[dict]) -> None:
    by_doc: dict[str, int] = {}
    for t in targets:
        by_doc[t["sourceDocument"]] = by_doc.get(t["sourceDocument"], 0) + 1
    print("=" * 78)
    print(f"Parsed {len(targets)} targets:", by_doc)
    print("=" * 78)
    for t in targets:
        flag = "  [CLEANED]" if t["_fixes"] else ""
        print(f"\n{t['id']:<10} | {t['sourceLabel']}{flag}")
        print(f"   {t['text']}")
        for bad, good in t["_fixes"]:
            print(f"      fix: {bad!r} -> {good!r}")

    # ID-collision check against existing targets
    existing = {t["id"] for t in json.loads(TARGETS_FILE.read_text())}
    collisions = sorted(t["id"] for t in targets if t["id"] in existing)
    print("\n" + "=" * 78)
    print(f"Existing targets: {len(existing)}  |  new: {len(targets)}  "
          f"|  total after append: {len(existing) + len(targets)}")
    print(f"ID collisions with existing data: {collisions or 'none'}")
    cleaned = sum(1 for t in targets if t["_fixes"])
    print(f"Rows with spacing fixes: {cleaned}  |  verbatim: {len(targets) - cleaned}")
    print(f"Countries: {sorted({t['country'] for t in targets})}")
    print("=" * 78)


def main() -> None:
    ap = argparse.ArgumentParser(description="Ingest Mongolia LDN targets")
    ap.add_argument("--append", action="store_true",
                    help="append parsed targets into mongolia-targets.json")
    args = ap.parse_args()

    targets = parse()
    qc_report(targets)

    # Strip the transient QC field before persisting.
    clean = [{k: v for k, v in t.items() if k != "_fixes"} for t in targets]

    PREVIEW_FILE.write_text(json.dumps(clean, indent=2, ensure_ascii=False) + "\n")
    print(f"\nPreview written: {PREVIEW_FILE.relative_to(REPO_ROOT)}")

    if args.append:
        existing = json.loads(TARGETS_FILE.read_text())
        existing_ids = {t["id"] for t in existing}
        dupes = [t["id"] for t in clean if t["id"] in existing_ids]
        if dupes:
            raise SystemExit(f"Refusing to append: ID collisions {dupes}")
        existing.extend(clean)
        TARGETS_FILE.write_text(json.dumps(existing, indent=2, ensure_ascii=False) + "\n")
        print(f"Appended {len(clean)} targets -> {TARGETS_FILE.relative_to(REPO_ROOT)} "
              f"(now {len(existing)})")
    else:
        print("\n(dry run; pass --append to write into mongolia-targets.json)")
        print(f"Provenance PDF for docProvenance: {SOURCE_PDF}")


if __name__ == "__main__":
    main()
