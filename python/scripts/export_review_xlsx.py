"""
Export an extraction result as an expert-review spreadsheet.

Takes a `*.extracted.json` file (the output of `python -m src.extract`) and
writes the "Targets for review" sheet the country-office experts already
know: one row per extracted target with its verbatim source text, extracted
activities, and blank Review decision / Reviewer comments columns (the
decision column carries the usual dropdown). This replaces the ad-hoc
generator used for the 2 Jul 2026 Sri Lanka NWRP/NFAP review, which never
lived in the repo.

Two columns exist so reviewers can interpret non-English and multilingual
documents correctly (lesson from that review round: a reviewer compared our
machine-translated Sinhala rows against the document's own English section
and reasonably concluded the text had been "re-written"):
  - "Language of source": the language block the row was extracted from.
  - "English text": "machine translation" when the working text was
    translated by the pipeline (textOriginalSource == "source"), otherwise
    "document wording".

Rows are written in input order; extract.py already emits document order
(see _sort_by_document_position), so the sheet reads top-to-bottom like the
document. Re-sorting here would fight deliberate manual edits, so it is not
attempted.

Usage (from python/):
    uv run python -m scripts.export_review_xlsx \
        --input path/to/document.extracted.json \
        --output path/to/targets_for_review.xlsx \
        --country "Sri Lanka" \
        --document "National Water Resources Policy of Sri Lanka (2023)" \
        --doc NWRP \
        --convention "Water"
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

HEADERS = [
    ("Country", 10),
    ("Target Title", 22),
    ("Target Text", 52),
    ("Document", 30),
    ("Doc", 8),
    ("Convention", 11),
    ("Extracted activities", 50),
    ("Source section", 13),
    ("Source text (original language)", 46),
    ("Extraction", 12),
    ("Language of source", 12),
    ("English text", 16),
    ("Review decision", 18),
    ("Reviewer comments", 34),
]

REVIEW_CHOICES = '"Approve as-is,Approve with edits,Needs discussion,Reject"'

LANG_NAMES = {
    "en": "English",
    "es": "Spanish",
    "fr": "French",
    "mn": "Mongolian",
    "si": "Sinhala",
    "ta": "Tamil",
    "und": "unidentified",
}


def _sections(item: dict) -> str:
    seen: list[str] = []
    for src in item.get("sources") or []:
        section = str(src.get("section", "")).strip()
        if section and section not in seen:
            seen.append(section)
    return "; ".join(seen)


def _source_text(item: dict) -> str:
    if item.get("textOriginal"):
        return str(item["textOriginal"])
    quotes = [
        str(src.get("sourceText", "")).strip()
        for src in item.get("sources") or []
        if str(src.get("sourceText", "")).strip()
    ]
    return "\n\n".join(quotes)


def _extraction_note(item: dict) -> str:
    note = str(item.get("textCleanup") or "")
    if item.get("_provenanceFlag"):
        note = f"{note}; FLAGGED" if note else "FLAGGED"
    return note


def build_review_sheet(
    items: list[dict],
    *,
    country: str,
    document: str,
    doc: str,
    convention: str,
) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Targets for review"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="0B4F8A")
    wrap = Alignment(wrap_text=True, vertical="top")

    for col, (name, width) in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = wrap
        ws.column_dimensions[get_column_letter(col)].width = width

    for row, item in enumerate(items, start=2):
        lang = str(item.get("language") or "en")
        translated = item.get("textOriginalSource") == "source"
        values = [
            country,
            item.get("label", ""),
            item.get("text", ""),
            document,
            doc,
            convention,
            item.get("activities", ""),
            _sections(item),
            _source_text(item),
            _extraction_note(item),
            LANG_NAMES.get(lang, lang),
            "machine translation" if translated else "document wording",
            "",
            "",
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = wrap

    decision_col = get_column_letter(len(HEADERS) - 1)
    dv = DataValidation(type="list", formula1=REVIEW_CHOICES, allow_blank=True)
    dv.add(f"{decision_col}2:{decision_col}{max(len(items) + 1, 2)}")
    ws.add_data_validation(dv)
    ws.freeze_panes = "C2"
    return wb


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", required=True, help="Extracted-targets JSON file")
    parser.add_argument("--output", required=True, help="Output .xlsx path")
    parser.add_argument("--country", required=True, help="Country display name")
    parser.add_argument("--document", required=True, help="Full document title")
    parser.add_argument("--doc", required=True, help="Short document code (e.g. NWRP)")
    parser.add_argument("--convention", default="", help="Convention column value")
    args = parser.parse_args()

    items = json.loads(Path(args.input).read_text())
    if not isinstance(items, list):
        raise SystemExit(f"{args.input} does not contain a JSON array of targets")

    wb = build_review_sheet(
        items,
        country=args.country,
        document=args.document,
        doc=args.doc,
        convention=args.convention,
    )
    out = Path(args.output)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"Wrote {len(items)} rows to {out}")


if __name__ == "__main__":
    main()
