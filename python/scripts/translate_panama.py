"""
Panama Target Translation Script — one-off Spanish → English.

Reads the Panama xlsx from the SharePoint sync, translates each row's
`Target Text` and `Target Name` from Spanish to English via Claude Opus 4.6
(through OpenRouter), applies R11.7 validation (language detection + numeric
diff + acronym whitelist), and writes three artefacts:

    python/data/panama-targets.json              — 284 translated targets
    python/data/panama-translation-flags.json    — rows flagged for review
    python/data/panama-translation-metadata.json — reproducibility sidecar

The script is committed to the repo as a reproducibility record, not a live
pipeline stage. It can be re-run if the xlsx updates. The disk cache in
`python/output/.cache/panama_translation/` keys translations by
(system+user+model), so a re-run with the same input is free and
deterministic.

Invocation:

    cd python && uv run python -m scripts.translate_panama

Requirements:
    - `OPENROUTER_API_KEY` (or `LLM_API_KEY`) in the project-root `.env`
    - SharePoint sync mounted at
      `dev_data_scripts/sharepoint_sync/Panama/data_panama_8Apr26 1.xlsx`
"""

from __future__ import annotations

import asyncio
import hashlib
import json
import logging
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from lingua import Language, LanguageDetectorBuilder
from openpyxl import load_workbook

from src.llm import call_llm_batch

# ─── Logging setup ──────────────────────────────────────────────────────────

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger("translate_panama")

# ─── Paths ──────────────────────────────────────────────────────────────────

PROJECT_ROOT = Path(__file__).resolve().parents[2]
XLSX_PATH = (
    PROJECT_ROOT
    / "dev_data_scripts"
    / "sharepoint_sync"
    / "Panama"
    / "data_panama_8Apr26 1.xlsx"
)
OUTPUT_DIR = PROJECT_ROOT / "python" / "data"
TARGETS_OUTPUT = OUTPUT_DIR / "panama-targets.json"
FLAGS_OUTPUT = OUTPUT_DIR / "panama-translation-flags.json"
METADATA_OUTPUT = OUTPUT_DIR / "panama-translation-metadata.json"

# ─── Translation config ─────────────────────────────────────────────────────

TRANSLATION_MODEL = "anthropic/claude-opus-4-6"
TRANSLATION_TEMPERATURE = 0.0
CACHE_NAMESPACE = "panama_translation"

# Acronyms that must survive translation verbatim. Match with word boundaries
# so short codes like "NP" don't false-positive on Spanish substrings
# (adversarial A8). The preflight at the start of main() drops whitelist
# entries with zero source hits so we don't chase ghosts.
ACRONYM_WHITELIST = [
    "REDD+",
    "NDC",
    "NBSAP",
    "NAP",
    "LDN",
    "BTR",
    "NR7",
    "IRMF",
    "SPGCF",
    "CNR",
    "UNFCCC",
    "CBD",
    "NP",
    "ENR",
    "GCF",
    "SDG",
]

TRANSLATION_SYSTEM = """You are a professional translator specialising in UN climate and biodiversity policy documents.
Translate the provided Spanish text into clear, faithful English suitable for policy analysis.

Hard rules:
- Preserve ALL numeric values and units exactly as written (percentages, hectares, years, monetary amounts, counts). Normalise decimal separators so Spanish-style "11,5%" becomes "11.5%". Do not add or drop any numbers.
- Preserve these acronyms VERBATIM in the English output: REDD+, NDC, NBSAP, NAP, LDN, BTR, NR7, IRMF, SPGCF, CNR, UNFCCC, CBD, NP, ENR, GCF, SDG. Never expand them, translate them, or change capitalisation.
- Keep structural elements intact: bullets, line breaks, numbered lists, and colons stay where they were.
- Return only the English translation. No commentary, no notes, no brackets explaining choices.
- If the input is already in English, return it unchanged.
- If the text is a mix of English and Spanish, translate the Spanish parts and leave the English parts untouched."""

TRANSLATION_USER = """Translate this text to English:

{text}"""

# ─── Language detection ─────────────────────────────────────────────────────

_DETECTOR = LanguageDetectorBuilder.from_languages(
    Language.ENGLISH, Language.SPANISH
).build()


def language_confidence(text: str, language: Language) -> float:
    """Return the detector's confidence (0-1) for `text` being in `language`."""
    if not text or not text.strip():
        return 0.0
    for result in _DETECTOR.compute_language_confidence_values(text):
        if result.language == language:
            return float(result.value)
    return 0.0


# ─── Numeric diffing ────────────────────────────────────────────────────────

# Match bare numbers like 1234, 12.5, 11,5 (Spanish decimal), 30%, 5.277 million
# We capture the full numeric token so we can normalise and compare.
_NUMERIC_RE = re.compile(r"(\d{1,3}(?:[.,]\d{3})+|\d+[.,]\d+|\d+)(%)?")


def extract_numbers(text: str) -> set[str]:
    """
    Pull normalised numeric tokens out of `text`.

    Decimals are normalised to a dot separator: both "11,5%" (Spanish) and
    "11.5%" (English) canonicalise to "11.5%". Thousand separators (longer
    groups) are stripped so "1,000" and "1.000" both canonicalise to "1000".
    A trailing "%" is preserved. Comparison is order-free because numbers
    can legitimately reorder during translation.

    Known limitation: a 3-digit decimal like "3.141" is ambiguous and will
    be read as thousand-separated (→ "3141"). Acceptable for policy targets,
    which typically round to 1-2 decimal places.
    """
    normalised: set[str] = set()
    for match in _NUMERIC_RE.finditer(text or ""):
        raw, pct = match.group(1), match.group(2) or ""
        # Decimal with 1-2 fraction digits (either separator): normalise to dot.
        if re.match(r"^\d+[.,]\d{1,2}$", raw):
            raw = raw.replace(",", ".")
        else:
            # Thousand separators or longer groups: strip to canonical integer.
            raw = raw.replace(".", "").replace(",", "")
        normalised.add(f"{raw}{pct}")
    return normalised


def diff_numbers(source: str, translated: str) -> list[str]:
    """Return human-readable descriptions of numeric tokens that drifted."""
    src_nums = extract_numbers(source)
    tgt_nums = extract_numbers(translated)
    missing = src_nums - tgt_nums
    extra = tgt_nums - src_nums
    issues: list[str] = []
    for n in sorted(missing):
        issues.append(f"number '{n}' missing from translation")
    for n in sorted(extra):
        issues.append(f"number '{n}' appears in translation but not source")
    return issues


# ─── Acronym preservation ───────────────────────────────────────────────────


def acronym_preflight(rows: list[dict[str, Any]]) -> list[str]:
    """
    Drop acronyms that never appear in the source. 30-second scan so we don't
    flag translations for missing acronyms that were never there.
    """
    all_source = " ".join(
        (r["text_es"] or "") + " " + (r["label_es"] or "") for r in rows
    )
    active = [
        a
        for a in ACRONYM_WHITELIST
        if re.search(rf"(?<!\w){re.escape(a)}(?!\w)", all_source, re.IGNORECASE)
    ]
    logger.info(
        f"Acronym preflight: {len(active)}/{len(ACRONYM_WHITELIST)} active — {active}"
    )
    return active


def check_acronyms(
    source: str,
    translated: str,
    active_acronyms: list[str],
) -> list[str]:
    """Flag any whitelist acronym that appears in source but not in translation."""
    issues: list[str] = []
    for acronym in active_acronyms:
        pattern = rf"(?<!\w){re.escape(acronym)}(?!\w)"
        in_source = bool(re.search(pattern, source, re.IGNORECASE))
        in_target = bool(re.search(pattern, translated))
        if in_source and not in_target:
            issues.append(f"acronym '{acronym}' not preserved in translation")
    return issues


# ─── Xlsx loading ───────────────────────────────────────────────────────────


def sha256_of_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def load_panama_rows() -> list[dict[str, Any]]:
    """Read the Panama xlsx and return a flat list of row dicts."""
    if not XLSX_PATH.exists():
        raise FileNotFoundError(
            f"Panama xlsx not found at {XLSX_PATH}. "
            "Is the SharePoint sync mounted?"
        )
    wb = load_workbook(XLSX_PATH, read_only=True)
    ws = wb["targets"]
    rows: list[dict[str, Any]] = []
    per_doc_counter: dict[str, int] = {}
    for raw in ws.iter_rows(min_row=2, values_only=True):
        if raw is None or not raw[1]:  # skip rows with no target text
            continue
        country, text_es, label_es, document, source, convention, doc, type_, _updated = (
            raw + (None,) * (9 - len(raw))
        )[:9]
        if not doc:
            logger.warning(f"Row with label '{label_es}' has no Doc value; skipping.")
            continue
        per_doc_counter[doc] = per_doc_counter.get(doc, 0) + 1
        target_id = f"panama_{doc}_{per_doc_counter[doc]}"
        rows.append(
            {
                "id": target_id,
                "country": country or "Panama",
                "text_es": str(text_es).strip(),
                "label_es": str(label_es or "").strip() or target_id,
                "document": str(document or "").strip(),
                "source_url": str(source or "").strip(),
                "convention": str(convention or "").strip(),
                "doc": str(doc).strip(),
                "type": str(type_ or "").strip(),
            }
        )
    logger.info(f"Loaded {len(rows)} Panama rows from xlsx.")
    return rows


# ─── Translation ────────────────────────────────────────────────────────────


async def translate_texts(texts: list[str]) -> list[str]:
    """Batch-translate via the shared LLM client. Uses disk cache."""
    calls = [
        {
            "system": TRANSLATION_SYSTEM,
            "user": TRANSLATION_USER.format(text=t),
            "model": TRANSLATION_MODEL,
            "temperature": TRANSLATION_TEMPERATURE,
        }
        for t in texts
    ]
    return await call_llm_batch(
        calls,
        cache_namespace=CACHE_NAMESPACE,
        desc="Translate Panama targets",
    )


# ─── Main ───────────────────────────────────────────────────────────────────


async def amain() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    rows = load_panama_rows()
    xlsx_sha = sha256_of_file(XLSX_PATH)
    logger.info(f"xlsx sha256: {xlsx_sha}")

    active_acronyms = acronym_preflight(rows)

    # Pre-check: language detection on source text. Flag mixed-language rows
    # BEFORE translating so reviewers see the source issue separately from any
    # translation drift.
    source_flags: dict[str, list[str]] = {}
    for row in rows:
        es_conf = language_confidence(row["text_es"], Language.SPANISH)
        if es_conf < 0.8:
            source_flags.setdefault(row["id"], []).append(
                f"source text Spanish confidence {es_conf:.2f} < 0.80 (mixed-language row)"
            )

    logger.info(
        f"Pre-translation: {len(source_flags)} row(s) flagged as mixed-language."
    )

    # Translate. Uses batched concurrent calls via the LLM client.
    logger.info(f"Translating {len(rows)} target texts with {TRANSLATION_MODEL}…")
    text_translations = await translate_texts([row["text_es"] for row in rows])

    logger.info(f"Translating {len(rows)} target labels…")
    label_translations = await translate_texts([row["label_es"] for row in rows])

    # Post-checks: for each row, run language, numeric, and acronym validation
    # on the translated text.
    all_flags: list[dict[str, Any]] = []
    targets: list[dict[str, Any]] = []

    for row, text_en, label_en in zip(rows, text_translations, label_translations, strict=True):
        text_en = (text_en or "").strip()
        label_en = (label_en or "").strip()

        reasons: list[str] = list(source_flags.get(row["id"], []))

        # Post-check language on translated text only (labels are too short for
        # reliable detection).
        if text_en:
            en_conf = language_confidence(text_en, Language.ENGLISH)
            if en_conf < 0.9:
                reasons.append(
                    f"translated text English confidence {en_conf:.2f} < 0.90 "
                    "(possible translation language drift)"
                )
        else:
            reasons.append("translated text is empty")

        # Numeric drift
        reasons.extend(diff_numbers(row["text_es"], text_en))

        # Acronym preservation
        reasons.extend(check_acronyms(row["text_es"], text_en, active_acronyms))
        reasons.extend(
            check_acronyms(row["label_es"], label_en, active_acronyms)
        )

        targets.append(
            {
                "id": row["id"],
                "text": text_en,
                "textOriginal": row["text_es"],
                "sourceDocument": row["doc"],
                "sourceLabel": label_en or row["label_es"],
                "sourceLabelOriginal": row["label_es"],
                "country": row["country"],
                "sourceUrl": row["source_url"],
            }
        )

        if reasons:
            all_flags.append(
                {
                    "id": row["id"],
                    "sourceDocument": row["doc"],
                    "sourceLabelOriginal": row["label_es"],
                    "textOriginal": row["text_es"],
                    "text": text_en,
                    "reasons": reasons,
                }
            )

    # Write outputs
    TARGETS_OUTPUT.write_text(
        json.dumps(targets, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    logger.info(f"Wrote {len(targets)} targets to {TARGETS_OUTPUT}")

    FLAGS_OUTPUT.write_text(
        json.dumps(all_flags, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    logger.info(f"Wrote {len(all_flags)} flagged rows to {FLAGS_OUTPUT}")

    metadata = {
        "model": TRANSLATION_MODEL,
        "temperature": TRANSLATION_TEMPERATURE,
        "promptSystem": TRANSLATION_SYSTEM,
        "promptUserTemplate": TRANSLATION_USER,
        "runDate": datetime.now(timezone.utc).isoformat(),
        "inputXlsxPath": str(XLSX_PATH.relative_to(PROJECT_ROOT)),
        "inputXlsxSha256": xlsx_sha,
        "totalRows": len(rows),
        "flaggedRows": len(all_flags),
        "acronymWhitelistActive": active_acronyms,
        "cacheNamespace": CACHE_NAMESPACE,
    }
    METADATA_OUTPUT.write_text(
        json.dumps(metadata, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    logger.info(f"Wrote metadata to {METADATA_OUTPUT}")

    logger.info(
        f"Done. {len(targets)} rows translated, {len(all_flags)} flagged. "
        f"See {FLAGS_OUTPUT.name} for review."
    )


def main() -> None:
    asyncio.run(amain())


if __name__ == "__main__":
    main()
